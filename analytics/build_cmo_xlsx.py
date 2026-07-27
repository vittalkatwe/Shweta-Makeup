#!/usr/bin/env python3
"""Build CMO-grade Creative Design Insights Excel workbook."""
import os, csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

OUT = 'analytics/Creative_Design_Insights_CMO.xlsx'
THUMB_DIR = 'analytics/creative_thumbnails'
TMP_DIR = 'analytics/_tmp_thumbs'
os.makedirs(TMP_DIR, exist_ok=True)

# Palette
NAVY = '1F2A44'; GOLD = 'D4A547'
GREEN = '4FAB6E'; GREEN_LT = 'E7F4EA'
RED = 'D9534F'; RED_LT = 'FBE5E4'
AMBER = 'F0AD4E'; AMBER_LT = 'FDF1E0'
GREY = '6C757D'; GREY_LT = 'F4F4F4'; WHITE = 'FFFFFF'

THIN = Side(border_style='thin', color='D0D0D0')
MED = Side(border_style='medium', color='B0B0B0')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_HDR = Border(left=MED, right=MED, top=MED, bottom=MED)

def fill(c): return PatternFill('solid', fgColor=c)
def font(size=10, bold=False, color='000000', italic=False):
    return Font(name='Calibri', size=size, bold=bold, color=color, italic=italic)
def align(h='left', v='center', wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def title(ws, cell, text, size=18, color=NAVY):
    ws[cell] = text
    ws[cell].font = font(size=size, bold=True, color=color)
    ws[cell].alignment = align('left', 'center', False)

def subtitle(ws, cell, text, size=10, color=GREY):
    ws[cell] = text
    ws[cell].font = font(size=size, italic=True, color=color)
    ws[cell].alignment = align('left', 'center', False)

def section(ws, cell, text, size=14, color=NAVY):
    ws[cell] = text
    ws[cell].font = font(size=size, bold=True, color=color)
    ws[cell].alignment = align('left', 'center', False)

def hdr_row(ws, row, cells, bg=NAVY, fg=WHITE):
    for i, val in enumerate(cells):
        c = ws.cell(row=row, column=2+i)
        c.value = val
        c.fill = fill(bg); c.font = font(size=10, bold=True, color=fg)
        c.alignment = align('center', 'center', True); c.border = BORDER_HDR

def data_row(ws, row, cells, bg=WHITE, bold_first=False, align_first='left'):
    for i, val in enumerate(cells):
        c = ws.cell(row=row, column=2+i)
        c.value = val
        c.fill = fill(bg); c.font = font(size=10, bold=(i==0 and bold_first), color='333333')
        c.alignment = align(align_first if i==0 else 'center', 'center', True)
        c.border = BORDER

# Load data
def num(v):
    if v in ('', None, '—'): return None
    try: return float(v)
    except: return None

rows = []
with open('analytics/creatives_joined.csv') as f:
    for r in csv.DictReader(f):
        rows.append({
            'ad': r['ad_name'], 'campaign': r['campaign'], 'adset': r['adset'],
            'bucket': r.get('bucket', ''),
            'spend': num(r['spend']), 'imp': num(r['impressions']), 'reach': num(r['reach']),
            'clicks': num(r['clicks']), 'ctr': num(r['ctr']), 'cpc': num(r['cpc']), 'cpm': num(r['cpm']),
            'mlpv': num(r['meta_lpv']), 'mic': num(r['meta_ic']), 'mpur': num(r['meta_pur']),
            'mp_hp': num(r['mp_homepage']), 'mp_pp': num(r['mp_paypage']),
            'mp_init': num(r['mp_init']), 'mp_succ': num(r['mp_success']),
            'cr': num(r['cr']), 'cpp': num(r['cost_per_pur']),
            'thumb': r.get('thumbnail', ''),
        })

by_ad = {r['ad']: r for r in rows}

def resize_thumb(ad_name, px=110):
    """Resize thumbnail to a fixed pixel size and return new path."""
    r = by_ad.get(ad_name)
    if not r or not r['thumb'] or not os.path.exists(r['thumb']):
        return None
    out = f'{TMP_DIR}/{os.path.basename(r["thumb"])}'
    img = PILImage.open(r['thumb']).convert('RGB')
    img.thumbnail((px, px), PILImage.LANCZOS)
    # pad to square
    new = PILImage.new('RGB', (px, px), (244, 244, 244))
    new.paste(img, ((px - img.width)//2, (px - img.height)//2))
    new.save(out)
    return out

def insert_image(ws, ad_name, anchor, px=110):
    path = resize_thumb(ad_name, px)
    if not path: return False
    img = XLImage(path)
    img.width = px; img.height = px
    img.anchor = anchor
    ws.add_image(img)
    return True

wb = Workbook()

# ============================================================
# SHEET 1: EXECUTIVE SUMMARY
# ============================================================
ws = wb.active; ws.title = '1. Executive Summary'
ws.sheet_view.showGridLines = False

ws.column_dimensions['A'].width = 2
ws.column_dimensions['B'].width = 6
ws.column_dimensions['C'].width = 42
ws.column_dimensions['D'].width = 95

title(ws, 'B2', 'Meta Ads — Creative Design Insights')
subtitle(ws, 'B3', 'Lifetime period: 17 Mar 2026 → 26 May 2026   ·   72 creatives   ·   ₹2,90,108 spend   ·   Account act_2526677564455677')
subtitle(ws, 'B4', 'Mixpanel funnel tracking installed 12 May 2026.   ·   187 backend-confirmed purchases (15 days)   ·   Meta-attributed lifetime: 960 (≈5× inflated by attribution)')
subtitle(ws, 'B5', 'Prepared 26 May 2026 · Source data: meta_ads_full_report + per_day_report + daily funnel snapshot')

# KPI tiles
ws.row_dimensions[8].height = 55
kpis = [
    ('Lifetime Spend', '₹2,90,108', NAVY),
    ('Backend Purchases', '187', GREEN),
    ('Avg Cost / Purchase', '₹1,551', AMBER),
    ('Days Active', '63 of 71', NAVY),
]
for i, (label, val, color) in enumerate(kpis):
    start = 2 + i*2
    c = ws.cell(row=8, column=start)
    c.value = f'{label}\n{val}'
    c.fill = fill(color); c.font = font(size=12, bold=True, color=WHITE)
    c.alignment = align('center', 'center', True); c.border = BORDER_HDR
    ws.merge_cells(start_row=8, start_column=start, end_row=8, end_column=start+1)

ws.row_dimensions[9].height = 55
kpis2 = [
    ('Unique Creatives', '72', NAVY),
    ('Total Impressions', '18.27 L', NAVY),
    ('Homepage Visitors', '4,650', GREEN),
    ('Overall CR (post-12 May)', '4.02%', GREEN),
]
for i, (label, val, color) in enumerate(kpis2):
    start = 2 + i*2
    c = ws.cell(row=9, column=start)
    c.value = f'{label}\n{val}'
    c.fill = fill(color); c.font = font(size=12, bold=True, color=WHITE)
    c.alignment = align('center', 'center', True); c.border = BORDER_HDR
    ws.merge_cells(start_row=9, start_column=start, end_row=9, end_column=start+1)

# Reset column widths back for findings section
ws.column_dimensions['B'].width = 6
for col in 'CDEFGHI':
    ws.column_dimensions[col].width = 0  # reset
ws.column_dimensions['C'].width = 42
ws.column_dimensions['D'].width = 95

section(ws, 'B12', '🎯 Eight headline findings')

findings = [
    ('1', 'Hairstyling beats makeup as the hook',
     'Every Mixpanel-confirmed Winner pitches hairstyling — 7 of 7. The single makeup-led test that landed (Skill-V5-Test, 8.1% CR) reframed makeup as an attention/identity outcome ("Catch Everyone\'s Eye"), not a skill claim. Pure-makeup creatives (2nd Set - Video2, eyeshadow close-up) only worked in the legacy pre-pivot funnel and cost 1.5× more per click.'),
    ('2', 'Losers have HIGHER CTR than Winners (3.23% vs 3.19%)',
     'CTR alone cannot predict purchase. Losers also pay 20% higher CPM (₹262 vs ₹219 — Meta penalizes irrelevant creatives via auction). The funnel break shows up at Pay-Page → Initiate (25% for Losers vs 60% for Winners) and at Initiate → Success (0% vs 80%). This is bottom-of-funnel rejection — a promise mismatch between ad and offer.'),
    ('3', 'Winning recipe is reproducible — a 7-for-7 pattern',
     'Hairstyling outcome shown + quantified promise ("10 mins" / "15 mins" / "5 days") + ₹299 price anchor on the creative + Hindi-English second-person headline + pink/coral branding + in-action photo (not stock). Skill-A72 hits all six and delivers 7.9% lifetime CR for ₹151 per purchase.'),
    ('4', 'Five anti-patterns to kill now (₹2,733 already wasted)',
     'Skill-A82 (lipstick shop products — off-theme), Skill-A21 (text-only yellow warning — no visual), Skill-A80 (sticky note on floor — looks like personal post), Skill-A74 (BREAKING news clickbait — high CTR, broken intent), General-C1 (generic salon stock — no specific outcome). All ≥₹500 spent, 0–1 backend purchases. Zero salvage value.'),
    ('5', '"–Winner" duplicates reset Meta\'s learning and underperform their originals',
     'Renaming a winning creative and relaunching as a fresh ad resets the learning phase. Skill-A63–Winner, Self-A30–Winner, Skill-A72-Winner, Skill-A78-Winner all delivered 0 purchases vs the originals\' 8 each. The "Winner" suffix is survivorship-bias signal, not durable performance. Promote by budget on the original ad ID — do not duplicate-and-rename.'),
    ('6', 'Daily mix is a 7× CR lever — bigger than creative choice alone',
     'Same creative pool delivered 0.83% CR on 18 May (₹1,557 per purchase) vs 5.85% CR on 21 May (₹225 per purchase). On the worst day, 47% of spend went to Winner-Campaign rename duplicates and the actual MP-confirmed winners (Skill-V8, Skill-A72, Skill-A63, Skill-A71, Self-A30) were either dormant or starved. The fix is operational, not creative.'),
    ('7', 'Skill-A72 is the standout — scale it 10×',
     'Lifetime 7.9% CR at ₹151 per purchase. Single-day record (24 May): 10.7% CR at ₹111 per purchase. Only ₹1,213 lifetime spend — the most underfunded winner in the portfolio. Reuse the same ad ID at 10× daily budget for 7 days before declaring fatigue. Highest leverage available right now.'),
    ('8', '92% of historical spend predates Mixpanel — treat Meta wins as directional',
     '₹2.66 lakh of ₹2.90 lakh ran before the 12 May tracking install. Sales Ad - A/B alone burned ₹1.05 lakh on 395 Meta-reported purchases — but Meta over-reports by 5× (960 Meta vs 187 backend lifetime). True return on Sales Ad - A/B likely 70–80 purchases at ~₹1,300 each, not the ₹265 per purchase Meta shows.'),
]

for i, (num_lbl, headline, detail) in enumerate(findings):
    r = 14 + i*2
    ws.row_dimensions[r].height = 78

    c = ws.cell(row=r, column=2)
    c.value = num_lbl
    c.fill = fill(GOLD); c.font = font(size=18, bold=True, color=WHITE)
    c.alignment = align('center', 'center', False); c.border = BORDER

    c = ws.cell(row=r, column=3)
    c.value = headline
    c.fill = fill(GREY_LT); c.font = font(size=11, bold=True, color=NAVY)
    c.alignment = align('left', 'center', True); c.border = BORDER

    c = ws.cell(row=r, column=4)
    c.value = detail
    c.fill = fill(WHITE); c.font = font(size=10, color='333333')
    c.alignment = align('left', 'center', True); c.border = BORDER

# Navigation
section(ws, 'B32', '📑 Workbook contents')
sheets_info = [
    ('1. Executive Summary', 'KPIs · eight headline findings'),
    ('2. Winners Playbook', 'Top 9 creatives with thumbnails · the winning recipe'),
    ('3. Losers to Kill', 'Anti-patterns to retire today · spend waste audit'),
    ('4. Funnel Diagnostics', 'Where each bucket breaks in the conversion funnel'),
    ('5. Metric Sensitivity', 'What CTR / CR / CPM / CPC each tell you'),
    ('6. Daily Mix Analysis', 'Best vs worst days · why daily allocation matters'),
    ('7. Action Plan', 'Concrete 30-day creative + ops moves'),
    ('8. All Creatives Database', 'Sortable list of all 72 ads with bucket labels'),
]
ws.row_dimensions[33].height = 6
for i, (name, desc) in enumerate(sheets_info):
    r = 34 + i
    ws.row_dimensions[r].height = 22
    c = ws.cell(row=r, column=2)
    c.value = '→'
    c.font = font(size=14, bold=True, color=GOLD)
    c.alignment = align('center', 'center', False)

    c = ws.cell(row=r, column=3)
    c.value = name
    c.font = font(size=11, bold=True, color=NAVY)
    c.alignment = align('left', 'center', False)
    c.fill = fill(GREY_LT); c.border = BORDER

    c = ws.cell(row=r, column=4)
    c.value = desc
    c.font = font(size=10, color='333333')
    c.alignment = align('left', 'center', False)
    c.fill = fill(WHITE); c.border = BORDER

ws.freeze_panes = 'A6'

wb.save(OUT)
print('✓ Sheet 1: Executive Summary')

# ============================================================
# SHEET 2: WINNERS PLAYBOOK
# ============================================================
ws = wb.create_sheet('2. Winners Playbook')
ws.sheet_view.showGridLines = False

ws.column_dimensions['A'].width = 2
ws.column_dimensions['B'].width = 18    # Thumbnail
ws.column_dimensions['C'].width = 22    # Ad name
ws.column_dimensions['D'].width = 12    # Backend Pur
ws.column_dimensions['E'].width = 10    # CR
ws.column_dimensions['F'].width = 12    # Cost/Pur
ws.column_dimensions['G'].width = 10    # Spend
ws.column_dimensions['H'].width = 10    # CTR
ws.column_dimensions['I'].width = 9    # CPC
ws.column_dimensions['J'].width = 9    # CPM
ws.column_dimensions['K'].width = 11    # HP visits
ws.column_dimensions['L'].width = 60    # Why it works

title(ws, 'B2', 'Winners Playbook')
subtitle(ws, 'B3', '10 top creatives, all hairstyling-led, all Mixpanel-confirmed. Together: 137 backend purchases on ₹40,102 spend ≈ ₹293 avg cost / purchase. Skill-A72 is the standout.')
subtitle(ws, 'B4', 'The visual recipe: hairstyling outcome + quantified promise + ₹299 anchor + Hindi-English headline + pink/coral branding + in-action photo (not stock).')

# Headers
ws.row_dimensions[6].height = 38
hdr_row(ws, 6, ['Creative', 'Ad Name', 'Backend Pur', 'CR %', 'Cost / Pur', 'Spend', 'CTR', 'CPC', 'CPM', 'Homepage Visits', 'Why it works (design read)'])

# Winners — ordered by backend purchases, including Meta-only top performers
winners = [
    ('Skill-V5-Test', 26, '8.1%', '₹275', '₹7,153', '2.72%', '₹11.8', '₹321', 320,
     'Tight face close-up on brown-skin model + headline "Start Makeup, Catch Everyone\'s Eye". The only makeup creative that worked — and only because it pitches makeup as identity/attention, not as a skill to learn. Pink/coral header. Highest single-creative volume of all post-12-May ads.'),
    ('SKUP1-WIN-V', 27, '3.7%', 'n/a', '₹10,353', '2.32%', '₹8.5', '₹197', 736,
     'Video creative — two-women conversation/interview setup. Story format, not headline-heavy. Highest absolute volume (27 backend pur), best at top-of-funnel reach (736 homepage visits). Most of its spend predates MP; lifetime CR shown is from utm_content match for the post-MP window only.'),
    ('Skill-V8– Winner', 15, '5.3%', '₹273', '₹4,101', '3.19%', '₹7.1', '₹226', 284,
     'Hands holding finished curls + "ONLY ₹299 Trending Hairstyles AT JUST ₹299" + pink CTA. Result + price + CTA on one frame. Twin of Skill-V8 — same creative scaled in Winner ad set.'),
    ('Skill-V8', 15, '6.7%', '₹156', '₹2,338', '3.74%', '₹4.9', '₹184', 225,
     'Same image as Winner version but in Testing ad set. Cheaper CPC because broader audience; better CR because of correct ad-set–creative match. Best ₹/purchase among high-volume creatives.'),
    ('Self-A24-WIN1', 12, '3.7%', '₹295', '₹3,539', '3.03%', '₹5.8', '₹174', 324,
     '"Love Hairstyling? But not sure where to start?" + multi-photo collage of hairstyles. Identity/question hook. Healthy volume but slightly lower CR than the result-shot family.'),
    ('Skill-A27-WIN1', 10, '2.6%', '₹455', '₹4,555', '2.65%', '₹7.4', '₹196', 382,
     '"Create such looks in just 15 mins" + finished bun shot + "Join Hairstyle Master Class". Quantified time promise. High homepage volume but lower conversion — could be audience-spread effect, watch ₹/pur.'),
    ('Skill-A71', 8, '5.2%', '₹289', '₹2,312', '3.09%', '₹6.7', '₹207', 154,
     '"Partitioning + Sectioning = Perfect Hairstyles" + split-screen showing technique steps. Names a SPECIFIC technique — under-tested category that could carry 4–5 more variants.'),
    ('Self-A30', 8, '6.0%', '₹288', '₹2,303', '3.79%', '₹10.5', '₹396', 134,
     '"Make Perfect Curls in 10 Minutes" + behind-the-scenes phone-and-ring-light setup. Aspirational + relatable home angle. Best CR among Self-* family. Watch CPC — single-day 24 May spiked to ₹14.78 (fatigue signal).'),
    ('Skill-A63', 8, '5.2%', '₹279', '₹2,235', '3.20%', '₹6.8', '₹219', 155,
     '"Apke Hairstyles main Finishing nahi hai?" + multi-style result grid + "Join 5-Day Hairstyle Class". Problem-aware Hindi second-person hook + outcome grid + duration callout. Textbook winner.'),
    ('Skill-A72', 8, '7.9%', '₹152', '₹1,213', '3.02%', '₹6.2', '₹188', 101,
     '★ STANDOUT ★ "Say Bye to Expensive Hairstyle Courses" + reviews screenshot + ₹299 anchor. Social proof + price + jealousy-of-competitors trifecta. Most efficient creative in the portfolio. Lifetime 7.9% CR, single-day 10.7%. Scale 10× immediately.'),
]

for i, (ad, succ, cr, cpp, spend, ctr, cpc, cpm, hp, why) in enumerate(winners):
    r = 8 + i*7  # 7 rows per winner: image takes ~6 rows
    # Row heights — first row is metric header
    for offset in range(6):
        ws.row_dimensions[r + offset].height = 17

    # Insert thumbnail at column B
    anchor_cell = f'B{r}'
    insert_image(ws, ad, anchor_cell, px=110)

    # Merge image area
    ws.merge_cells(start_row=r, start_column=2, end_row=r+5, end_column=2)

    # Ad name (merged vertically)
    c = ws.cell(row=r, column=3)
    c.value = ad
    if ad == 'Skill-A72':
        c.fill = fill(GOLD)
        c.font = font(size=12, bold=True, color=WHITE)
    else:
        c.fill = fill(GREEN_LT); c.font = font(size=12, bold=True, color=NAVY)
    c.alignment = align('center', 'center', True); c.border = BORDER
    ws.merge_cells(start_row=r, start_column=3, end_row=r+5, end_column=3)

    # Metrics
    metrics = [succ, cr, cpp, f'₹{spend}', ctr, cpc, cpm, hp]
    for j, v in enumerate(metrics):
        c = ws.cell(row=r, column=4+j)
        c.value = v
        c.fill = fill(GREEN_LT if ad != 'Skill-A72' else AMBER_LT)
        c.font = font(size=11, bold=True, color=NAVY)
        c.alignment = align('center', 'center', False); c.border = BORDER
        ws.merge_cells(start_row=r, start_column=4+j, end_row=r+5, end_column=4+j)

    # Why it works (merged)
    c = ws.cell(row=r, column=12)
    c.value = why
    c.fill = fill(WHITE); c.font = font(size=10, color='333333')
    c.alignment = align('left', 'center', True); c.border = BORDER
    ws.merge_cells(start_row=r, start_column=12, end_row=r+5, end_column=12)

ws.freeze_panes = 'A7'
wb.save(OUT)
print('✓ Sheet 2: Winners Playbook')

# ============================================================
# SHEET 3: LOSERS TO KILL
# ============================================================
ws = wb.create_sheet('3. Losers to Kill')
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 2
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 11
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 9
ws.column_dimensions['G'].width = 12
ws.column_dimensions['H'].width = 11
ws.column_dimensions['I'].width = 11
ws.column_dimensions['J'].width = 13
ws.column_dimensions['K'].width = 65

title(ws, 'B2', 'Losers to Kill — Today')
subtitle(ws, 'B3', '7 Mixpanel-confirmed money losers. ₹4,023 spent · 3 backend purchases · 0.07% CR. Cut now to redirect spend.')
subtitle(ws, 'B4', 'Pattern: high or normal CTR, healthy homepage visits, then catastrophic drop at the payment page. The break is in the promise, not the hook.')

ws.row_dimensions[6].height = 38
hdr_row(ws, 6, ['Creative', 'Ad Name', 'Spend', 'CTR', 'CPC', 'Homepage Visits', 'Pay-Page', 'Backend Pur', 'Failure Mode', 'Why it fails (design read)'], bg=RED, fg=WHITE)

losers = [
    ('Self-A34', '₹857', '4.18%', '₹9.6', 60, 4, 1, 'Soft hook',
     '"Match Curls in Buns at home" with woman + practice doll. Closer to a winner aesthetically but the hook is vague — "Match Curls in Buns" isn\'t specific enough. ₹857/purchase is 5× the winner average. Cap or kill.'),
    ('Self-A31', '₹557', '3.15%', '₹4.4', 66, 6, 1, 'Bold but no result',
     '"Don\'t Go to Beauty Parlour for Hairstyles" + dark salon door image. Bold copy but no aspirational outcome shown — visitor doesn\'t see what success looks like. Functional copy without visual reward.'),
    ('General-C1', '₹551', '2.95%', '₹13.4', 22, 3, 0, 'Generic / stock',
     'Top-down salon stock photo + "5 Reasons to join our Hairstyle MasterClass". No specific outcome, no visual hook. CTR is below median. Highest CPC in the loser pool — Meta finds this irrelevant. Zero salvage.'),
    ('Skill-A55', '₹541', '3.23%', '₹9.0', 31, 4, 0, 'Buried headline',
     'Multi-hairstyle collage with the headline ("Want to Learn Celebrity Hairstyle Techniques?") small and at the top. Visuals look promising but the hook is hard to read on mobile. Worth rewriting headline with specific outcome + price, then re-test.'),
    ('Skill-A21', '₹515', '2.09%', '₹8.9', 30, 0, 0, 'Text-only warning',
     'Yellow warning-icon text block with no real image. Looks like a forum/listing screenshot, not an ad. 0 visitors made it to the payment page. Won\'t scale at any spend. Kill, do not iterate.'),
    ('Skill-A74', '₹501', '5.08%', '₹2.7', 113, 8, 1, 'CTR-trap / clickbait',
     '"BREAKING news" red tabloid banner. Highest CTR of any loser AND suspiciously cheap clicks (₹2.65 CPC). Meta is finding curiosity-clickers, not learners. 113 visits → 1 purchase. Classic clickbait — kill regardless of CTR optics.'),
    ('Skill-A52', '₹501', '4.13%', '₹6.3', 39, 4, 0, 'Process not outcome',
     '"Still Practising on Dummy?" + photo of beginner practicing on a doll. Names the pain but shows an unfinished state — visitor doesn\'t see what success looks like. Pair this hook with a finished-style image and retest.'),
]

for i, (ad, spend, ctr, cpc, hp, pp, pur, mode, why) in enumerate(losers):
    r = 8 + i*7
    for offset in range(6):
        ws.row_dimensions[r + offset].height = 17

    insert_image(ws, ad, f'B{r}', px=110)
    ws.merge_cells(start_row=r, start_column=2, end_row=r+5, end_column=2)

    c = ws.cell(row=r, column=3); c.value = ad
    c.fill = fill(RED_LT); c.font = font(size=12, bold=True, color=RED)
    c.alignment = align('center', 'center', True); c.border = BORDER
    ws.merge_cells(start_row=r, start_column=3, end_row=r+5, end_column=3)

    metrics = [spend, ctr, cpc, hp, pp, pur]
    for j, v in enumerate(metrics):
        c = ws.cell(row=r, column=4+j); c.value = v
        c.fill = fill(RED_LT); c.font = font(size=11, bold=True, color=NAVY)
        c.alignment = align('center', 'center', False); c.border = BORDER
        ws.merge_cells(start_row=r, start_column=4+j, end_row=r+5, end_column=4+j)

    c = ws.cell(row=r, column=10); c.value = mode
    c.fill = fill(RED); c.font = font(size=10, bold=True, color=WHITE)
    c.alignment = align('center', 'center', True); c.border = BORDER
    ws.merge_cells(start_row=r, start_column=10, end_row=r+5, end_column=10)

    c = ws.cell(row=r, column=11); c.value = why
    c.fill = fill(WHITE); c.font = font(size=10, color='333333')
    c.alignment = align('left', 'center', True); c.border = BORDER
    ws.merge_cells(start_row=r, start_column=11, end_row=r+5, end_column=11)

# Total burn summary
total_row = 8 + len(losers)*7 + 1
ws.row_dimensions[total_row].height = 32
c = ws.cell(row=total_row, column=2); c.value = 'TOTAL WASTE'
c.fill = fill(NAVY); c.font = font(size=12, bold=True, color=WHITE)
c.alignment = align('center', 'center', False); c.border = BORDER_HDR
ws.merge_cells(start_row=total_row, start_column=2, end_row=total_row, end_column=3)

burn_cells = ['₹4,023', '3.23%', '₹8.87', 361, 29, 3, 'Multiple', '₹1,341 per purchase (8.6× winner avg of ₹262)']
for j, v in enumerate(burn_cells):
    c = ws.cell(row=total_row, column=4+j); c.value = v
    c.fill = fill(NAVY); c.font = font(size=11, bold=True, color=WHITE)
    c.alignment = align('center', 'center', True); c.border = BORDER_HDR

ws.freeze_panes = 'A7'
wb.save(OUT)
print('✓ Sheet 3: Losers to Kill')

# ============================================================
# SHEET 4: FUNNEL DIAGNOSTICS
# ============================================================
ws = wb.create_sheet('4. Funnel Diagnostics')
ws.sheet_view.showGridLines = False

for col, w in [('A', 2), ('B', 32), ('C', 16), ('D', 16), ('E', 16), ('F', 70)]:
    ws.column_dimensions[col].width = w

title(ws, 'B2', 'Funnel Diagnostics')
subtitle(ws, 'B3', 'Where the buckets diverge in the conversion funnel. Click and reach the homepage similarly — but Losers reject the offer at the payment page.')

# Bucket performance table
section(ws, 'B6', 'Bucket-level metrics (Mixpanel ground truth)')

ws.row_dimensions[8].height = 38
hdr_row(ws, 8, ['Bucket', 'Ads', 'Spend', 'Median CTR', 'Median CPC', 'Median CPM', 'Backend Purchases', 'Median CR'])

bucket_data = [
    ('WINNERS (MP-confirmed)', '7', '₹21,655', '3.19%', '₹6.84', '₹219', '88', '5.97%', GREEN_LT),
    ('MID-TIER (MP-confirmed)', '23', '₹16,542', '2.66%', '₹8.52', '₹232', '40', '3.70%', AMBER_LT),
    ('LOSERS (MP-confirmed)', '7', '₹4,023', '3.23%', '₹8.87', '₹262', '3', '0.00%', RED_LT),
    ('WINNERS (Meta-only, pre-12-May)', '6', '₹1,78,742', '2.74%', '₹6.02', '₹179', '654 (Meta)', 'n/a — pre-MP', GREEN_LT),
    ('PRE-MP (rest, untracked)', '23', '₹58,196', '3.21%', '₹7.48', '₹239', '163 (Meta)', 'n/a — pre-MP', GREY_LT),
    ('LEAD-GEN (no LPV tracking)', '4', '₹6,915', '2.29%', '₹2.93', '₹68', '0', 'n/a — no LP', GREY_LT),
]

for col_idx, w in enumerate([32, 8, 14, 13, 13, 13, 18, 16]):
    ws.column_dimensions[chr(ord('B')+col_idx)].width = w

for i, row_data in enumerate(bucket_data):
    r = 9 + i
    ws.row_dimensions[r].height = 24
    bg = row_data[-1]
    for j, val in enumerate(row_data[:-1]):
        c = ws.cell(row=r, column=2+j)
        c.value = val
        c.fill = fill(bg); c.font = font(size=10, color='333333', bold=(j==0))
        c.alignment = align('left' if j==0 else 'center', 'center', True)
        c.border = BORDER

# Funnel diagnostic
section(ws, 'B17', 'Funnel-stage performance — where each bucket loses')

# Reset column widths for this section
ws.column_dimensions['B'].width = 32
ws.column_dimensions['C'].width = 16
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 16
ws.column_dimensions['F'].width = 70

ws.row_dimensions[19].height = 38
hdr_row(ws, 19, ['Funnel Stage', 'WINNERS', 'MID-TIER', 'LOSERS', 'Read'])

funnel = [
    ('Homepage → Pay Page', '11.0%', '8.3%', '9.1%', 'Top-of-funnel similar across buckets — clicks behave normally for all three groups'),
    ('Pay Page → Initiate', '60.0%', '50.0%', '25.0%', '⚠ Loser break #1 — visitors reach the page but reject the offer (promise mismatch)'),
    ('Initiate → Success', '80.0%', '71.4%', '0.0%', '⚠ Loser break #2 — even those who start checkout abandon at payment'),
    ('Net Homepage → Success (CR)', '5.97%', '3.70%', '0.00%', 'Loss compounds at the bottom of the funnel — design fix is alignment, not hook'),
]

for i, row_data in enumerate(funnel):
    r = 20 + i
    ws.row_dimensions[r].height = 32
    for j, val in enumerate(row_data):
        c = ws.cell(row=r, column=2+j)
        c.value = val
        if j == 0:
            c.fill = fill(GREY_LT); c.font = font(size=11, bold=True, color=NAVY)
        elif j == 1:
            c.fill = fill(GREEN_LT); c.font = font(size=12, bold=True, color=GREEN)
        elif j == 2:
            c.fill = fill(AMBER_LT); c.font = font(size=12, bold=True, color='C77B07')
        elif j == 3:
            c.fill = fill(RED_LT); c.font = font(size=12, bold=True, color=RED)
        else:
            c.fill = fill(WHITE); c.font = font(size=10, color='333333')
        c.alignment = align('left' if j in (0,4) else 'center', 'center', True)
        c.border = BORDER

# Interpretation callout
ws.row_dimensions[25].height = 12
section(ws, 'B26', '💡 The Insight')

ws.row_dimensions[27].height = 80
c = ws.cell(row=27, column=2)
c.value = '⚡'
c.fill = fill(GOLD); c.font = font(size=24, bold=True, color=WHITE)
c.alignment = align('center', 'center', False); c.border = BORDER

c = ws.cell(row=27, column=3)
c.value = (
    'Losers and Winners look nearly identical at the top of the funnel — same CTR, same homepage-arrival rate. '
    'The divergence happens AFTER the visitor lands on the site. Loser creatives promise something (cheap products, '
    'BREAKING news, irrelevant problems) that the landing page does not deliver — so visitors bounce at the payment page. '
    '\n\nThis is a CREATIVE-TO-OFFER mismatch problem, not a creative-quality problem. The fix is to ensure every creative\'s '
    'promise (price + outcome + format) matches what the landing page sells. Kill creatives where the on-image promise '
    'diverges from the landing page; do not optimize them.'
)
c.fill = fill(GREY_LT); c.font = font(size=11, color='333333')
c.alignment = align('left', 'center', True); c.border = BORDER
ws.merge_cells(start_row=27, start_column=3, end_row=27, end_column=6)

ws.freeze_panes = 'A6'
wb.save(OUT)
print('✓ Sheet 4: Funnel Diagnostics')

# ============================================================
# SHEET 5: METRIC SENSITIVITY
# ============================================================
ws = wb.create_sheet('5. Metric Sensitivity')
ws.sheet_view.showGridLines = False
for col, w in [('A', 2), ('B', 18), ('C', 24), ('D', 70), ('E', 50)]:
    ws.column_dimensions[col].width = w

title(ws, 'B2', 'Metric Sensitivity — what each KPI actually tells you')
subtitle(ws, 'B3', 'Reading individual metrics in isolation is misleading. This sheet shows how each metric correlates (or fails to correlate) with revenue.')

ws.row_dimensions[6].height = 38
hdr_row(ws, 6, ['Metric', 'What it measures', 'What the data shows', 'How to read it'])

metric_rows = [
    ('CTR', 'Strength of the hook (top-of-funnel attention)',
     'Winners 3.19% · Losers 3.23%. The two highest-CTR creatives in the dataset (Skill-A74 5.08%, Skill-A82 4.25%) are both Losers. CTR alone is uncorrelated with backend conversion.',
     'Necessary but not sufficient. Use CTR ≥ 2.5% as a baseline filter, then judge on PP→Initiate and final CR. A creative with high CTR + low PP→Initiate is pulling curiosity-clickers, not buyers.'),
    ('CPC', 'How expensive each click is',
     'Winners ₹6.84 · Losers ₹8.87 (similar). But the outliers tell the real story: Skill-A74 at ₹2.65 CPC (suspiciously cheap, 1 purchase on 113 visits) and lead-gen campaigns at ₹2.93 CPC (0 visibility into conversion).',
     'Cheap clicks are suspect when they\'re not accompanied by funnel data. If CPC drops below ₹4 on a new creative, expect lower-intent audience. Verify with PP→Initiate within 24 hours of launch.'),
    ('CPM', 'How much Meta charges to serve 1,000 impressions',
     'Losers ₹262 · Winners ₹219 — Losers cost 20% more to deliver. The pre-MP top spenders had ₹125–₹206 CPM. When CPM creeps above ₹250 on a new ad, it\'s often a leading indicator that funnel conversion will also be weak.',
     'CPM is the cleanest "Meta hates this creative" signal — Meta\'s auction penalizes low engagement (low video watch %, low reactions). Treat CPM > ₹260 as a warning before MP data arrives.'),
    ('CR % (Mixpanel)', 'True homepage-to-purchase conversion rate',
     'Winners 5.97% · Mid 3.70% · Losers 0.00%. Sitewide average is 4.02%. This is the only metric that strictly correlates with revenue — Mixpanel fires only on successful Razorpay charge.',
     'Source of truth. Anything below 3% needs investigation; below 1% needs killing. Always read alongside funnel-stage data to know WHERE the break is (top-funnel weakness vs offer mismatch).'),
    ('Cost / Purchase', 'Spend efficiency on the only metric that matters',
     'Winners cluster ₹150–₹300 (best: Skill-A72 ₹151, Skill-V8 ₹156). Acceptable upper bound ₹400; above ₹500 is a kill signal. Loser median ₹1,341 — 8.6× winner average.',
     'The bottom-line metric. Combine with CR% to identify "expensive winner" (high spend, high CR) vs "scaling opportunity" (low spend, high CR — e.g. Skill-A72). The latter is where to allocate next.'),
    ('Backend Purchases', 'Absolute volume of confirmed sales',
     'Top 3: SKUP1-WIN-V (27, Meta-attributed), Skill-V5-Test (26), Skill-V8 + Skill-V8–Winner (30 combined). Volume scales with spend, so always read it as a ratio with spend, not in isolation.',
     'Use for scale planning, not for ranking. Always normalize by spend (₹/purchase) or visitors (CR%) before comparing creatives.'),
]

for i, (metric, what, data, read) in enumerate(metric_rows):
    r = 8 + i*2
    ws.row_dimensions[r].height = 95

    c = ws.cell(row=r, column=2); c.value = metric
    c.fill = fill(GOLD); c.font = font(size=14, bold=True, color=WHITE)
    c.alignment = align('center', 'center', True); c.border = BORDER

    c = ws.cell(row=r, column=3); c.value = what
    c.fill = fill(GREY_LT); c.font = font(size=10, color=NAVY, bold=True)
    c.alignment = align('left', 'center', True); c.border = BORDER

    c = ws.cell(row=r, column=4); c.value = data
    c.fill = fill(WHITE); c.font = font(size=10, color='333333')
    c.alignment = align('left', 'center', True); c.border = BORDER

    c = ws.cell(row=r, column=5); c.value = read
    c.fill = fill(GREEN_LT); c.font = font(size=10, color='333333', italic=True)
    c.alignment = align('left', 'center', True); c.border = BORDER

ws.freeze_panes = 'A6'
wb.save(OUT)
print('✓ Sheet 5: Metric Sensitivity')

# ============================================================
# SHEET 6: DAILY MIX ANALYSIS
# ============================================================
ws = wb.create_sheet('6. Daily Mix Analysis')
ws.sheet_view.showGridLines = False
for col, w in [('A', 2), ('B', 14), ('C', 9), ('D', 9), ('E', 10), ('F', 13), ('G', 65)]:
    ws.column_dimensions[col].width = w

title(ws, 'B2', 'Daily Mix Analysis — the 7× CR lever')
subtitle(ws, 'B3', 'Same creative pool can deliver 0.83% CR or 5.85% CR depending on WHICH ads are funded that day. Source: meta_ads_per_day_report.xlsx')

ws.row_dimensions[6].height = 38
hdr_row(ws, 6, ['Date', 'Spend', 'Ads Live', 'CR %', '₹ / Purchase', 'What was running that day'])

days = [
    ('21 May (best)', '₹2,477', 13, '5.85%', '₹225',
     '✅ Top spenders = pure MP-confirmed winners only: Skill-V8, Self-A30, Skill-A71, Skill-A63, Self-A34, Skill-A51. No Winner-rename duplicates. Clean focused mix.', GREEN_LT),
    ('23 May', '₹4,468', 19, '5.11%', '₹263',
     'Winner family + Skill-A72 + Skill-V8 dominant in spend. Some bleed to mid-tier but core winners protected.', GREEN_LT),
    ('13 May', '₹6,287', 16, '5.10%', '₹233',
     'Heavy concentration in legacy proven creatives: SKUP1-WIN-V (₹1,613, 12 Meta pur), SUB1-TEST-V (₹1,210, 3 MP), Skill-A27-WIN1 (₹1,087). Pre-newer-Skill-A* era — old guard at scale.', GREEN_LT),
    ('24 May', '₹5,145', 20, '4.08%', '₹303',
     'Skill-A72 had its 10.7% CR day here. But many filler ads diluted overall — 5 zero-conversion ads ate ₹1,328.', AMBER_LT),
    ('22 May', '₹4,742', 18, '2.91%', '₹474',
     '⚠ Spend spread too thin across 18 ads. ₹1,900+ went to General-C1 (0 conv), Skill-A57/A59/A2 (0 conv each), and three "–Winner" duplicates (low conv). Real winners diluted to ~₹275 each.', AMBER_LT),
    ('25 May', '₹5,123', 17, '2.95%', '₹342',
     'Similar pattern to 22 May — overspread across too many ads. The 5 known winners didn\'t get enough share.', AMBER_LT),
    ('18 May (worst)', '₹1,557', 12, '0.83%', '₹1,557',
     '❌ 47% of spend on Winner Campaign rename duplicates: SKUP1-WIN-V (₹298), Self-A24-WIN1 (₹284), Skill-A27-WIN1 (₹157). None of the newer high-CR creatives (Skill-A63, A71, A72, V8) running. Only Skill-V5-Test got 1 purchase.', RED_LT),
]

for i, (date, spend, ads, cr, cpp, mix, bg) in enumerate(days):
    r = 8 + i
    ws.row_dimensions[r].height = 56
    cells = [date, spend, ads, cr, cpp, mix]
    for j, v in enumerate(cells):
        c = ws.cell(row=r, column=2+j); c.value = v
        c.fill = fill(bg)
        if j == 0:
            c.font = font(size=11, bold=True, color=NAVY)
        elif j == 5:
            c.font = font(size=10, color='333333')
        else:
            c.font = font(size=11, bold=True, color=NAVY)
        c.alignment = align('left' if j == 5 else 'center', 'center', True)
        c.border = BORDER

# Implication
section(ws, 'B18', '💡 The operational fix (no new creatives needed)')
ws.row_dimensions[20].height = 130
c = ws.cell(row=20, column=2)
c.value = (
    'On 22 May, ₹4,742 spent across 18 ads delivered 10 purchases at ₹474 each (2.91% CR).\n\n'
    'If the same ₹4,742 had concentrated on the 5 known MP winners (Skill-V8 + Self-A30 + Skill-A72 + Skill-A63 + Skill-A71) '
    'at ~₹950 each — projected based on their proven cost-per-purchase of ₹150–₹290 — you would expect ~22–25 purchases at '
    '₹190–220 each.\n\n'
    '→ Same spend.  → 2.4× more purchases.  → 60% cost reduction.\n'
    '→ Zero new creatives required. Pure allocation discipline.\n\n'
    'The operational rule: at least 60% of daily spend must go to the top 5 MP-confirmed winners. '
    'Any new creative gets a ₹300 cap; if no backend purchase by ₹300, kill same day.'
)
c.fill = fill(GOLD); c.font = font(size=11, bold=False, color=WHITE)
c.alignment = align('left', 'center', True); c.border = BORDER
ws.merge_cells(start_row=20, start_column=2, end_row=20, end_column=7)

ws.freeze_panes = 'A6'
wb.save(OUT)
print('✓ Sheet 6: Daily Mix Analysis')

# ============================================================
# SHEET 7: ACTION PLAN
# ============================================================
ws = wb.create_sheet('7. Action Plan')
ws.sheet_view.showGridLines = False
for col, w in [('A', 2), ('B', 4), ('C', 13), ('D', 35), ('E', 60), ('F', 18), ('G', 14)]:
    ws.column_dimensions[col].width = w

title(ws, 'B2', '30-day Action Plan')
subtitle(ws, 'B3', 'Three categories: scale immediately, test in production, kill today. Each item has an owner timeline and a measurable success metric.')

ws.row_dimensions[6].height = 38
hdr_row(ws, 6, ['#', 'Priority', 'Action', 'Detail', 'Timeline', 'Success metric'])

actions = [
    # SCALE
    (1, 'SCALE', 'Scale Skill-A72 10× on existing ad ID',
     'Reuse same ad ID, push daily budget from ~₹40 to ~₹400 over 5 days. Do not duplicate-and-rename. Watch CR daily; expect fatigue around day 7–10.',
     'Week 1', 'CR ≥ 6%, CPP ≤ ₹250', GREEN_LT),
    (2, 'SCALE', 'Protect daily spend on the 5 MP-confirmed winners',
     'Skill-V8, Skill-V8–Winner, Skill-A72, Skill-A63, Skill-A71, Self-A30. At least 60% of daily ad spend must hit this set. Hard rule for buyer/agency.',
     'Immediate', 'Daily CR ≥ 4% on 9 of 10 days', GREEN_LT),
    (3, 'SCALE', 'Move Self-A24-WIN1 and Skill-A27-WIN1 to capped scale',
     '12 and 10 purchases respectively at slightly higher CPP (₹295, ₹455). Worth maintaining but cap at ₹500/day each to avoid budget drift.',
     'Week 1', 'CPP stays ≤ ₹500', GREEN_LT),
    # TEST
    (4, 'TEST', 'Build 4–5 named-technique creatives',
     'Skill-A71 (Partitioning) and Skill-A51 (Gota Patti) proved that specific-technique hooks work. Build variants for: Juda, Khopa, Messy Bun, Front-Puff, Reception Hairstyle.',
     'Week 2', '≥ 2 hit CR ≥ 5%', AMBER_LT),
    (5, 'TEST', 'Replicate Skill-A72 / Skill-V8 trifecta with 3 visual variants',
     'Same headline formula (₹299 + duration + outcome) on 3 different hairstyle photos / models. Hold copy and price constant; isolate the visual variable.',
     'Week 2', '≥ 1 matches Skill-A72 CR', AMBER_LT),
    (6, 'TEST', 'One controlled makeup creative to validate "makeup is dead"',
     'Same composition + same headline + ₹299 anchor as Skill-V8 but with finished-makeup look instead of hairstyle. If it doesn\'t beat 4% CR in 7 days, retire makeup as a hook permanently.',
     'Week 3', 'CR ≥ 4% OR retire makeup', AMBER_LT),
    # KILL
    (7, 'KILL', 'Pause 5 anti-pattern creatives today',
     'Skill-A82 (lipstick shop), Skill-A21 (text-only warning), Skill-A80 (sticky note), Skill-A74 (BREAKING news), General-C1 (generic salon). Combined ₹2,733 wasted with 0–1 purchases.',
     'Today', '₹0 future spend on these', RED_LT),
    (8, 'KILL', 'Stop creating "–Winner" rename duplicates',
     'Promote ads by budget on the original ad ID. Update naming convention with the buyer/agency. Right now you cannot distinguish true creative fatigue from naming-induced learning resets.',
     'Today', '0 new –Winner ads', RED_LT),
    (9, 'KILL', 'Cap or pause Bollywood/news-clipping creatives',
     'Skill-A68, Skill-A78, Skill-A55 — moderate CTR but unstable conversion. Cap at ₹500/week each; if no purchase at the cap, pause.',
     'Week 1', 'Each ad ≥ 1 pur per ₹500', RED_LT),
    # OPS
    (10, 'OPS', 'Instrument lead-gen campaigns or stop them',
     'Interest ad, Non Interest ad, 1k Interest ad, 1k Non Interest ad have NO LPV tracking — they spent ₹6,915 with no funnel visibility. Add LP instrumentation or pause.',
     'Week 1', 'Funnel visible OR ads off', GREY_LT),
    (11, 'OPS', 'Audit price callouts on every creative',
     'Skill-A79 shows ₹399 on the creative while live charge is ₹299. Promise mismatch will hurt PP→Initiate. Standardize price text on every running creative.',
     'Week 1', '100% creatives match live price', GREY_LT),
    (12, 'OPS', 'Fix utm_content normalization (en-dash vs hyphen)',
     'Skill-V8–Winner uses an en-dash in Meta but a hyphen in utm_content. Reports normalize, but downstream tools may under-count this creative.',
     'Week 2', 'No utm normalization required downstream', GREY_LT),
]

for i, (num, prio, action, detail, timeline, metric, bg) in enumerate(actions):
    r = 8 + i
    ws.row_dimensions[r].height = 55

    c = ws.cell(row=r, column=2); c.value = num
    c.fill = fill(NAVY); c.font = font(size=11, bold=True, color=WHITE)
    c.alignment = align('center', 'center', False); c.border = BORDER

    c = ws.cell(row=r, column=3); c.value = prio
    prio_color = {'SCALE': GREEN, 'TEST': AMBER, 'KILL': RED, 'OPS': GREY}[prio]
    c.fill = fill(prio_color); c.font = font(size=11, bold=True, color=WHITE)
    c.alignment = align('center', 'center', True); c.border = BORDER

    c = ws.cell(row=r, column=4); c.value = action
    c.fill = fill(bg); c.font = font(size=11, bold=True, color=NAVY)
    c.alignment = align('left', 'center', True); c.border = BORDER

    c = ws.cell(row=r, column=5); c.value = detail
    c.fill = fill(WHITE); c.font = font(size=10, color='333333')
    c.alignment = align('left', 'center', True); c.border = BORDER

    c = ws.cell(row=r, column=6); c.value = timeline
    c.fill = fill(bg); c.font = font(size=10, bold=True, color=NAVY)
    c.alignment = align('center', 'center', True); c.border = BORDER

    c = ws.cell(row=r, column=7); c.value = metric
    c.fill = fill(bg); c.font = font(size=10, color='333333', italic=True)
    c.alignment = align('center', 'center', True); c.border = BORDER

ws.freeze_panes = 'A7'
wb.save(OUT)
print('✓ Sheet 7: Action Plan')

# ============================================================
# SHEET 8: ALL CREATIVES DATABASE
# ============================================================
ws = wb.create_sheet('8. All Creatives DB')
ws.sheet_view.showGridLines = False

cols = ['Ad Name', 'Bucket', 'Campaign', 'Ad Set', 'Spend (₹)', 'Imp', 'Clicks',
        'CTR %', 'CPC (₹)', 'CPM (₹)', 'Meta LPV', 'Meta IC', 'Meta Pur',
        'MP Homepage', 'MP Pay-Page', 'MP Init', 'MP Success', 'CR %', 'Cost/Pur (₹)']
widths = [26, 18, 22, 22, 12, 10, 10, 9, 10, 10, 11, 11, 11, 13, 13, 11, 13, 9, 13]

ws.column_dimensions['A'].width = 2
for i, w in enumerate(widths):
    ws.column_dimensions[chr(ord('B')+i)].width = w

title(ws, 'B2', 'All 72 Creatives — sortable database')
subtitle(ws, 'B3', 'Lifetime data 17 Mar → 26 May. Color = bucket. Right-click column → Sort & Filter to drill in.')

ws.row_dimensions[5].height = 50
for i, h in enumerate(cols):
    c = ws.cell(row=5, column=2+i); c.value = h
    c.fill = fill(NAVY); c.font = font(size=10, bold=True, color=WHITE)
    c.alignment = align('center', 'center', True); c.border = BORDER_HDR

# Sort by spend desc
rows_sorted = sorted(rows, key=lambda x: -(x['spend'] or 0))

bucket_color = {
    'WINNER': GREEN_LT, 'WINNER_META_ONLY': GREEN_LT,
    'LOSER': RED_LT, 'LOSER_META_ONLY': RED_LT,
    'MID': AMBER_LT, 'PRE_MP': GREY_LT, 'LEADGEN': GREY_LT,
}

def fmt_pct(v): return f'{v*100:.2f}%' if v is not None else '—'
def fmt_num(v): return f'{v:,.0f}' if v is not None else '—'
def fmt_money(v): return f'{v:,.0f}' if v is not None else '—'
def fmt_money2(v): return f'{v:,.2f}' if v is not None else '—'

for idx, r in enumerate(rows_sorted):
    row = 6 + idx
    ws.row_dimensions[row].height = 20
    bg = bucket_color.get(r['bucket'], GREY_LT)
    vals = [
        r['ad'], r['bucket'], r['campaign'] or '—', r['adset'] or '—',
        fmt_money(r['spend']), fmt_num(r['imp']), fmt_num(r['clicks']),
        fmt_pct(r['ctr']), fmt_money2(r['cpc']), fmt_money(r['cpm']),
        fmt_num(r['mlpv']), fmt_num(r['mic']), fmt_num(r['mpur']),
        fmt_num(r['mp_hp']), fmt_num(r['mp_pp']), fmt_num(r['mp_init']),
        fmt_num(r['mp_succ']), fmt_pct(r['cr']), fmt_money(r['cpp']),
    ]
    for j, v in enumerate(vals):
        c = ws.cell(row=row, column=2+j); c.value = v
        c.fill = fill(bg)
        c.font = font(size=9, bold=(j==0), color=NAVY if j==0 else '333333')
        c.alignment = align('left' if j<=3 else 'right', 'center', False)
        c.border = BORDER

# Add autofilter
ws.auto_filter.ref = f'B5:{chr(ord("B")+len(cols)-1)}{5+len(rows_sorted)}'
ws.freeze_panes = 'C6'

wb.save(OUT)
print('✓ Sheet 8: All Creatives Database')

# Bring Executive Summary back to front
wb.active = wb['1. Executive Summary']
wb.save(OUT)

# Report file size
sz_kb = os.path.getsize(OUT) / 1024
print(f'\n📊 Workbook saved: {OUT} ({sz_kb:.0f} KB)')
print(f'   Sheets: {len(wb.sheetnames)} → {wb.sheetnames}')
