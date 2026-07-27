#!/usr/bin/env python3
"""Append Designer Playbook + Media Buyer Rules sheets to the CMO workbook."""
import os, csv, io
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

XLSX = 'analytics/Creative_Design_Insights_CMO.xlsx'

# Palette (matches existing workbook)
NAVY = '1F2A44'; GOLD = 'D4A547'
GREEN = '4FAB6E'; GREEN_LT = 'E7F4EA'; GREEN_DK = '2D7849'
RED = 'D9534F'; RED_LT = 'FBE5E4'; RED_DK = 'A03430'
AMBER = 'F0AD4E'; AMBER_LT = 'FDF1E0'
GREY = '6C757D'; GREY_LT = 'F4F4F4'; WHITE = 'FFFFFF'
PINK = 'EC8AAE'

THIN = Side(border_style='thin', color='D0D0D0')
MED = Side(border_style='medium', color='B0B0B0')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_HDR = Border(left=MED, right=MED, top=MED, bottom=MED)

def fill(c): return PatternFill('solid', fgColor=c)
def font(size=10, bold=False, color='000000', italic=False):
    return Font(name='Calibri', size=size, bold=bold, color=color, italic=italic)
def align(h='left', v='center', wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def title(ws, cell, text, size=18, color=NAVY):
    ws[cell] = text
    ws[cell].font = font(size=size, bold=True, color=color)
    ws[cell].alignment = align('left', 'center', False)

def subtitle(ws, cell, text, size=10, color=GREY):
    ws[cell] = text
    ws[cell].font = font(size=size, italic=True, color=color)
    ws[cell].alignment = align('left', 'center', False)

def section(ws, cell, text, size=14, color=NAVY):
    ws[cell] = text
    ws[cell].font = font(size=size, bold=True, color=color)
    ws[cell].alignment = align('left', 'center', False)

def insert_thumb(ws, thumb_path, anchor, px=120):
    """Resize thumbnail in-memory and embed without tmp files."""
    if not thumb_path or not os.path.exists(thumb_path):
        return False
    src = PILImage.open(thumb_path).convert('RGB')
    src.thumbnail((px, px), PILImage.LANCZOS)
    canvas = PILImage.new('RGB', (px, px), (244, 244, 244))
    canvas.paste(src, ((px - src.width)//2, (px - src.height)//2))
    buf = io.BytesIO()
    canvas.save(buf, format='PNG')
    buf.seek(0)
    img = XLImage(buf)
    img.width = px; img.height = px
    img.anchor = anchor
    ws.add_image(img)
    return True

# Load existing workbook
wb = load_workbook(XLSX)
print(f'Loaded {XLSX} with {len(wb.sheetnames)} sheets')

# Remove old playbook sheets if they exist (idempotent)
for old in ['9. Designer Playbook', '10. Media Buyer Rules']:
    if old in wb.sheetnames:
        del wb[old]

# ============================================================
# SHEET 9: DESIGNER PLAYBOOK
# ============================================================
ws = wb.create_sheet('9. Designer Playbook')
ws.sheet_view.showGridLines = False

# Column widths
ws.column_dimensions['A'].width = 2
ws.column_dimensions['B'].width = 4    # Number badge
ws.column_dimensions['C'].width = 18   # Winner thumb
ws.column_dimensions['D'].width = 38   # Winner notes
ws.column_dimensions['E'].width = 18   # Loser thumb
ws.column_dimensions['F'].width = 38   # Loser notes
ws.column_dimensions['G'].width = 30   # Takeaway

title(ws, 'B2', 'Designer Playbook — How to make ads that win')
subtitle(ws, 'B3', 'Plain-language guide for the creative designer. Every rule is grounded in real ad data from the last 70 days, ₹2.9 lakh spend, 187 confirmed sales.')
subtitle(ws, 'B4', 'Goal of every new creative: LOW cost per 1,000 views (CPM < ₹220) + HIGH click-through (CTR ≥ 3%) + HIGH conversion (CR ≥ 5%). Hit all three = a winner.')

# Winner profile callout
ws.row_dimensions[6].height = 8
section(ws, 'B7', '★ The target signature — what every new creative should aim for')

ws.row_dimensions[9].height = 95
c = ws.cell(row=9, column=2); c.value = '🎯'
c.fill = fill(GOLD); c.font = font(size=24, color=WHITE)
c.alignment = align('center', 'center', False); c.border = BORDER

c = ws.cell(row=9, column=3); c.value = (
    'CPM under ₹220\n(cheap impressions,\nbut not too cheap)'
)
c.fill = fill(GREEN_LT); c.font = font(size=11, bold=True, color=GREEN_DK)
c.alignment = align('center', 'center', True); c.border = BORDER

c = ws.cell(row=9, column=4); c.value = (
    'CTR 3% or higher\n(strong hook,\npeople click)'
)
c.fill = fill(GREEN_LT); c.font = font(size=11, bold=True, color=GREEN_DK)
c.alignment = align('center', 'center', True); c.border = BORDER

c = ws.cell(row=9, column=5); c.value = (
    'CR 5% or higher\n(once they land,\nthey buy)'
)
c.fill = fill(GREEN_LT); c.font = font(size=11, bold=True, color=GREEN_DK)
c.alignment = align('center', 'center', True); c.border = BORDER

c = ws.cell(row=9, column=6); c.value = (
    'Cost per sale\n< ₹250\n(profitable to scale)'
)
c.fill = fill(GREEN_LT); c.font = font(size=11, bold=True, color=GREEN_DK)
c.alignment = align('center', 'center', True); c.border = BORDER

c = ws.cell(row=9, column=7); c.value = (
    'Real examples that hit all 4:\nSkill-A72, Skill-V8,\nSkill-A71, Skill-A63'
)
c.fill = fill(GOLD); c.font = font(size=10, bold=True, color=WHITE)
c.alignment = align('center', 'center', True); c.border = BORDER

# Side-by-side comparisons
ws.row_dimensions[11].height = 12
section(ws, 'B12', '🔍 Side-by-side: same idea, different result')
subtitle(ws, 'B13', 'Each row shows a winning creative next to a losing creative from the same category. Same audience, same time. The visual is the only thing that changed.')

ws.row_dimensions[15].height = 32
headers = ['', 'Winner thumbnail', 'Why this WON ✓', 'Loser thumbnail', 'Why this LOST ✗', 'Takeaway for designer']
for i, h in enumerate(headers):
    c = ws.cell(row=15, column=2+i); c.value = h
    c.fill = fill(NAVY); c.font = font(size=10, bold=True, color=WHITE)
    c.alignment = align('center', 'center', True); c.border = BORDER_HDR

# Pair rows — load thumbnails from data folder
pairs = [
    {
        'n': 1,
        'win_ad': 'Skill-A72', 'win_thumb': 'analytics/creative_thumbnails/Skill-A72.png',
        'win_stat': '₹188 CPM · 3.0% CTR · 7.9% CR · ₹151 per sale',
        'win_why': (
            'Headline: "Say Bye to Expensive Hairstyle Courses"\n'
            '+ Reviews screenshot showing real students\n'
            '+ ₹299 price visible on the image\n'
            '+ Pink/red brand banner top and bottom\n\n'
            'Reads as: a competitor-killer offer with social proof and a clear price.'
        ),
        'lose_ad': 'Skill-A21', 'lose_thumb': 'analytics/creative_thumbnails/Skill-A21.png',
        'lose_stat': '₹185 CPM · 2.1% CTR · 0 sales',
        'lose_why': (
            'Just a yellow warning box with text.\n'
            'No image. No model. No result shown.\n'
            'Looks like a forum post or system error.\n\n'
            'Reads as: not a real ad — easy to scroll past.'
        ),
        'takeaway': 'A SCREENSHOT of reviews or comments beats a TEXT BOX every time. Show real people reacting, not just rules in a box.'
    },
    {
        'n': 2,
        'win_ad': 'Skill-V8', 'win_thumb': 'analytics/creative_thumbnails/Skill-V8.png',
        'win_stat': '₹184 CPM · 3.7% CTR · 6.7% CR · ₹156 per sale',
        'win_why': (
            'A hand holding a finished curled hairstyle.\n'
            '+ Text: "Trending Hairstyles AT JUST ₹299"\n'
            '+ Pink CTA button visible\n'
            '+ Dark background makes the curls pop\n\n'
            'Reads as: a clear finished result + a price the viewer can act on right now.'
        ),
        'lose_ad': 'Skill-A82', 'lose_thumb': 'analytics/creative_thumbnails/Skill-A82.png',
        'lose_stat': '₹128 CPM · 4.3% CTR · 0 sales',
        'lose_why': (
            'A shop display showing lipstick products.\n'
            'Nothing about hairstyles.\n'
            'Looks like a Bollywood beauty product post.\n\n'
            'People click expecting cheap lipsticks. They leave when the page sells a course.'
        ),
        'takeaway': 'Cheap CPM is NOT a good thing if the click is for the wrong reason. Always show the actual thing you sell — hairstyles, not lipsticks.'
    },
    {
        'n': 3,
        'win_ad': 'Self-A30', 'win_thumb': 'analytics/creative_thumbnails/Self-A30.png',
        'win_stat': '₹396 CPM · 3.8% CTR · 6.0% CR · ₹288 per sale',
        'win_why': (
            'A real "behind the scenes" photo:\n'
            'a phone on a ring light, stylist working on a model.\n'
            '+ Text: "Make Perfect Curls in 10 Minutes"\n'
            '+ ₹299 corner badge\n\n'
            'Reads as: this is a real working stylist. You can do this too.'
        ),
        'lose_ad': 'Skill-A80', 'lose_thumb': 'analytics/creative_thumbnails/Skill-A80.png',
        'lose_stat': '₹206 CPM · 4.1% CTR · 0 sales',
        'lose_why': (
            'A handwritten sticky note on a wooden floor.\n'
            'No model. No hairstyle. No price.\n'
            'Looks like a personal post, not a paid ad.\n\n'
            'People stop out of curiosity but leave because there is no offer.'
        ),
        'takeaway': '"Behind the scenes" works ONLY if the result and the price are also visible. Curiosity without a CTA = wasted clicks.'
    },
    {
        'n': 4,
        'win_ad': 'Skill-A63', 'win_thumb': 'analytics/creative_thumbnails/Skill-A63.png',
        'win_stat': '₹219 CPM · 3.2% CTR · 5.2% CR · ₹279 per sale',
        'win_why': (
            'Headline in Hindi-English mix:\n'
            '"Apke Hairstyles main Finishing nahi hai?"\n'
            '+ Grid of 4 finished hairstyle photos\n'
            '+ "Join 5-Day Hairstyle Class" CTA\n\n'
            'Reads as: it speaks the audience\'s problem AND shows the solution.'
        ),
        'lose_ad': 'General-C1', 'lose_thumb': 'analytics/creative_thumbnails/General-C1.png',
        'lose_stat': '₹397 CPM · 3.0% CTR · 0 sales',
        'lose_why': (
            'A stock-photo style top-down salon shot.\n'
            '+ "5 Reasons to join our Hairstyle MasterClass"\n'
            'No specific outcome. No specific learner.\n\n'
            'Reads as: a generic course ad. Nothing to grab onto.'
        ),
        'takeaway': 'A SPECIFIC PROBLEM ("Finishing nahi hai?") + a SHOWN RESULT (4 finished styles) beats generic salon photos and "5 Reasons" lists.'
    },
    {
        'n': 5,
        'win_ad': 'Skill-A71', 'win_thumb': 'analytics/creative_thumbnails/Skill-A71.png',
        'win_stat': '₹207 CPM · 3.1% CTR · 5.2% CR · ₹289 per sale',
        'win_why': (
            'A split image:\n'
            'left side shows partitioning, right side shows the final hairstyle.\n'
            '+ Text: "Partitioning + Sectioning = Perfect Hairstyles"\n'
            '+ "Join Hairstyle Class" button\n\n'
            'Reads as: shows you exactly which technique you will learn AND what it gives you.'
        ),
        'lose_ad': 'Skill-A55', 'lose_thumb': 'analytics/creative_thumbnails/Skill-A55.png',
        'lose_stat': '₹291 CPM · 3.2% CTR · 0 sales',
        'lose_why': (
            'A collage of hairstyles — visually similar to the winner.\n'
            'BUT the headline ("Want to Learn Celebrity Hairstyle Techniques?")\n'
            'is small and at the top, hard to read on mobile.\n'
            'No price. No specific technique named.\n\n'
            'Reads as: pretty pictures, no clear offer.'
        ),
        'takeaway': 'NAME the specific technique (Partitioning, Sectioning, Curls, Gota Patti). And put the headline BIG. Small headlines on collages get ignored.'
    },
    {
        'n': 6,
        'win_ad': 'Skill-V5-Test', 'win_thumb': 'analytics/creative_thumbnails/Skill-V5-Test.png',
        'win_stat': '₹321 CPM · 2.7% CTR · 8.1% CR · ₹275 per sale',
        'win_why': (
            'Tight close-up of a model\'s face.\n'
            '+ Text: "Start Makeup, Catch Everyone\'s Eye"\n'
            '+ Pink/coral branding\n\n'
            'The ONLY makeup angle that worked — because it pitches makeup as ATTENTION and IDENTITY, not a "skill to learn".'
        ),
        'lose_ad': 'Skill-A74', 'lose_thumb': 'analytics/creative_thumbnails/Skill-A74.png',
        'lose_stat': '₹135 CPM · 5.1% CTR · 1 sale',
        'lose_why': (
            'A red "BREAKING news" banner with images underneath.\n'
            'Looks like a tabloid news clipping.\n'
            'Cheap impressions, lots of clicks — but Meta is showing it to news-readers, not buyers.\n\n'
            '113 people visited the site. 1 bought.'
        ),
        'takeaway': 'NEVER USE NEWS/TABLOID STYLES. Cheap CPM + high CTR + no sales is the worst combo — you pay for clicks from people who never buy. Stick to clean branded layouts.'
    },
]

row = 16
for p in pairs:
    ws.row_dimensions[row].height = 130

    # Pair number badge
    c = ws.cell(row=row, column=2); c.value = p['n']
    c.fill = fill(GOLD); c.font = font(size=20, bold=True, color=WHITE)
    c.alignment = align('center', 'center', False); c.border = BORDER

    # Winner thumb cell — embed image
    insert_thumb(ws, p['win_thumb'], f'C{row}', px=120)
    c = ws.cell(row=row, column=3)
    c.fill = fill(GREEN_LT); c.border = BORDER

    # Winner notes
    c = ws.cell(row=row, column=4)
    c.value = f"{p['win_ad']}\n{p['win_stat']}\n\n{p['win_why']}"
    c.fill = fill(GREEN_LT); c.font = font(size=10, color=NAVY)
    c.alignment = align('left', 'top', True); c.border = BORDER

    # Loser thumb cell
    insert_thumb(ws, p['lose_thumb'], f'E{row}', px=120)
    c = ws.cell(row=row, column=5)
    c.fill = fill(RED_LT); c.border = BORDER

    # Loser notes
    c = ws.cell(row=row, column=6)
    c.value = f"{p['lose_ad']}\n{p['lose_stat']}\n\n{p['lose_why']}"
    c.fill = fill(RED_LT); c.font = font(size=10, color=NAVY)
    c.alignment = align('left', 'top', True); c.border = BORDER

    # Takeaway
    c = ws.cell(row=row, column=7)
    c.value = p['takeaway']
    c.fill = fill(GOLD); c.font = font(size=11, bold=True, color=WHITE)
    c.alignment = align('left', 'center', True); c.border = BORDER

    row += 1

# DO list
ws.row_dimensions[row].height = 12
row += 1
section(ws, f'B{row}', '✅ The DO list — 10 design moves that produce winners')
row += 2

dos = [
    ('Show a FINISHED hairstyle', 'Curls, buns, partitions, finished updos. The viewer must see the result they will get. Examples: Skill-V8, Skill-A63, Skill-A27-WIN1.'),
    ('Put the PRICE on the image', 'The ₹299 callout in a visible corner or banner. Every winner has this. Examples: Skill-A72, Skill-V8.'),
    ('Use a SPECIFIC TIME or DURATION', 'Not "easy" — write "10 mins", "15 mins", "5-day class". Quantified promises convert better. Examples: Self-A30 (10 mins), Skill-A27-WIN1 (15 mins), Skill-A63 (5-day).'),
    ('Name the SPECIFIC TECHNIQUE', 'Partitioning, Sectioning, Gota Patti, Front-Puff. Generic "hairstyling" doesn\'t convert. Examples: Skill-A71, Skill-A51.'),
    ('Speak the customer\'s PROBLEM in Hindi-English mix', 'Second-person voice. "Apke Hairstyles main Finishing nahi hai?" beats "Learn hairstyling now". Example: Skill-A63.'),
    ('Use PINK or CORAL brand banding', 'Top or bottom of the image. Makes it scannable as the same brand across creatives. Most winners do this.'),
    ('Show a REAL person or model — never stock', 'Behind-the-scenes shots (ring light, working stylist) feel real and convert. Example: Self-A30.'),
    ('Include a CTA BUTTON in the design', 'Visible "Join Hairstyle Class" or similar. Makes it look actionable. Examples: Skill-V8, Skill-A63.'),
    ('Use SOCIAL PROOF when you can', 'Screenshots of reviews, comments, "Clients Love..." — borrows credibility. Examples: Skill-A72 (review screenshots), Skill-A51 ("Clients Love Gota Patti").'),
    ('When using video, START with the RESULT', 'The first frame is what Meta uses as the thumbnail. Lead with a finished hairstyle, not the process or intro card.'),
]

for i, (rule, detail) in enumerate(dos):
    ws.row_dimensions[row].height = 40
    c = ws.cell(row=row, column=2); c.value = f'{i+1}'
    c.fill = fill(GREEN); c.font = font(size=13, bold=True, color=WHITE)
    c.alignment = align('center', 'center', False); c.border = BORDER

    c = ws.cell(row=row, column=3); c.value = rule
    c.fill = fill(GREEN_LT); c.font = font(size=11, bold=True, color=GREEN_DK)
    c.alignment = align('left', 'center', True); c.border = BORDER
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)

    c = ws.cell(row=row, column=5); c.value = detail
    c.fill = fill(WHITE); c.font = font(size=10, color='333333')
    c.alignment = align('left', 'center', True); c.border = BORDER
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
    row += 1

# DON'T list
ws.row_dimensions[row].height = 12
row += 1
section(ws, f'B{row}', '❌ The DON\'T list — 10 design choices that burn money')
row += 2

donts = [
    ('NEVER use news/tabloid styling', 'Red "BREAKING news" banners, newspaper layouts. They get cheap clicks from news-readers who never buy. Example of money loss: Skill-A74 (₹501 spent, 1 sale).'),
    ('NEVER show off-theme products', 'Lipsticks, makeup products, beauty store shelves. You sell a hairstyling COURSE, not products. Example: Skill-A82 (₹302 spent, 0 sales).'),
    ('NEVER use plain text-on-color boxes', 'Yellow warning boxes, error messages, forum-style text. They look like spam. Example: Skill-A21 (₹515 spent, 0 sales).'),
    ('NEVER use sticky-note or personal-post styles', 'Handwritten notes on floors, casual snapshots without structure. Looks organic but doesn\'t convert. Example: Skill-A80 (₹321 spent, 0 sales).'),
    ('NEVER use generic salon stock photos', 'Top-down salon shots, generic styling setups. Zero specificity = zero conversion. Example: General-C1 (₹551 spent, 0 sales).'),
    ('NEVER bury the headline in small text', 'If the headline is unreadable on mobile at thumb-size, the ad fails. Example: Skill-A55 (₹541 spent, 0 sales).'),
    ('NEVER show "practising" / unfinished work as the main image', 'Practice on a dummy, mid-process shots without result. Pain without aspiration = doesn\'t sell. Example: Skill-A52 (₹501 spent, 0 sales).'),
    ('NEVER show a model without a hairstyle outcome', 'Just a portrait isn\'t enough. The viewer must see what they will be able to DO. Example: Skill-A2 (₹492 spent, 1 sale).'),
    ('NEVER rename a winning ad and re-upload it', 'Calling it "—Winner" resets Meta\'s learning. The duplicates underperformed the originals every time. Example: Skill-A63–Winner, Self-A30–Winner.'),
    ('NEVER assume cheap CPM means good creative', 'Skill-A74 had ₹135 CPM (very cheap) and zero conversion. Cheap impressions from news-style ads = wrong audience. Always check sales, not just CPM.'),
]

for i, (rule, detail) in enumerate(donts):
    ws.row_dimensions[row].height = 40
    c = ws.cell(row=row, column=2); c.value = f'{i+1}'
    c.fill = fill(RED); c.font = font(size=13, bold=True, color=WHITE)
    c.alignment = align('center', 'center', False); c.border = BORDER

    c = ws.cell(row=row, column=3); c.value = rule
    c.fill = fill(RED_LT); c.font = font(size=11, bold=True, color=RED_DK)
    c.alignment = align('left', 'center', True); c.border = BORDER
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)

    c = ws.cell(row=row, column=5); c.value = detail
    c.fill = fill(WHITE); c.font = font(size=10, color='333333')
    c.alignment = align('left', 'center', True); c.border = BORDER
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
    row += 1

# Creative brief template
ws.row_dimensions[row].height = 12
row += 1
section(ws, f'B{row}', '📝 Creative brief template — designer fills this in BEFORE designing')
row += 2

brief_fields = [
    ('1. Hairstyle topic / technique', '(e.g. "Bun with Gota Patti", "Front-puff for parties", "Curls in 10 mins")'),
    ('2. Specific time promise', '(e.g. "10 minutes", "15 mins", "5-day class") — must be a number'),
    ('3. Price callout to put on image', '₹299 (verify with media buyer — must match the live website price)'),
    ('4. Main visual', '(e.g. "finished bun close-up", "split-shot: technique on left, result on right", "behind-the-scenes ring light setup")'),
    ('5. Headline in Hindi-English mix', '(e.g. "Apke ___ main Finishing nahi hai? — 5-Day Class")'),
    ('6. Audience problem this addresses', '(e.g. "stylists losing clients to better-trained competitors", "homemakers wanting a side income", "MUAs who can\'t do hair")'),
    ('7. Social proof element (if any)', '(e.g. review screenshot, "Clients love ___", testimonial quote, count of students)'),
    ('8. CTA text on the creative', '"Join Hairstyle Class" / "Start Now" / "₹299 Only"'),
    ('9. Brand colors used', 'Pink/coral banding (top, bottom, or both)'),
    ('10. Format (still / carousel / video)', '(If video: which winning still does the first frame mimic?)'),
]

ws.row_dimensions[row].height = 28
c = ws.cell(row=row, column=2); c.value = 'Field'
c.fill = fill(NAVY); c.font = font(size=10, bold=True, color=WHITE)
c.alignment = align('center', 'center', True); c.border = BORDER_HDR
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
c = ws.cell(row=row, column=5); c.value = 'Designer fills in here'
c.fill = fill(NAVY); c.font = font(size=10, bold=True, color=WHITE)
c.alignment = align('center', 'center', True); c.border = BORDER_HDR
ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
row += 1

for field, hint in brief_fields:
    ws.row_dimensions[row].height = 36
    c = ws.cell(row=row, column=2); c.value = field
    c.fill = fill(GREY_LT); c.font = font(size=10, bold=True, color=NAVY)
    c.alignment = align('left', 'center', True); c.border = BORDER
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)

    c = ws.cell(row=row, column=5); c.value = hint
    c.fill = fill(WHITE); c.font = font(size=10, color=GREY, italic=True)
    c.alignment = align('left', 'center', True); c.border = BORDER
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
    row += 1

# Pre-launch QA gate
ws.row_dimensions[row].height = 12
row += 1
section(ws, f'B{row}', '🚦 Pre-launch QA gate — every creative must pass ALL 10 before going live')
row += 2

qa = [
    'Does the image show a FINISHED hairstyle outcome?',
    'Is the PRICE (₹299) visible on the image?',
    'Is there a SPECIFIC NUMBER (10 mins, 15 mins, 5-day) on the image?',
    'Is the headline LARGE and readable on a phone at thumbnail size?',
    'Does it speak to the customer\'s PROBLEM in their language (Hindi-English mix OK)?',
    'Is the brand PINK/CORAL banding present somewhere?',
    'Is there a visible CTA ("Join Hairstyle Class" or similar)?',
    'Does the price callout on the image MATCH the live website price?',
    'Is the visual a real person/result (NOT a stock photo, NOT a news clipping, NOT a sticky note)?',
    'If video: does the first frame look like a winning still (because Meta uses it as the thumbnail)?',
]

ws.row_dimensions[row].height = 28
c = ws.cell(row=row, column=2); c.value = '#'
c.fill = fill(NAVY); c.font = font(size=10, bold=True, color=WHITE)
c.alignment = align('center', 'center', True); c.border = BORDER_HDR
c = ws.cell(row=row, column=3); c.value = 'Checklist question'
c.fill = fill(NAVY); c.font = font(size=10, bold=True, color=WHITE)
c.alignment = align('left', 'center', True); c.border = BORDER_HDR
ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
c = ws.cell(row=row, column=6); c.value = 'Y / N'
c.fill = fill(NAVY); c.font = font(size=10, bold=True, color=WHITE)
c.alignment = align('center', 'center', True); c.border = BORDER_HDR
c = ws.cell(row=row, column=7); c.value = 'If NO — fix and re-check'
c.fill = fill(NAVY); c.font = font(size=10, bold=True, color=WHITE)
c.alignment = align('left', 'center', True); c.border = BORDER_HDR
row += 1

for i, q in enumerate(qa):
    ws.row_dimensions[row].height = 30
    c = ws.cell(row=row, column=2); c.value = i+1
    c.fill = fill(GREY_LT); c.font = font(size=10, bold=True, color=NAVY)
    c.alignment = align('center', 'center', False); c.border = BORDER
    c = ws.cell(row=row, column=3); c.value = q
    c.fill = fill(WHITE); c.font = font(size=10, color='333333')
    c.alignment = align('left', 'center', True); c.border = BORDER
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
    c = ws.cell(row=row, column=6); c.value = ''
    c.fill = fill(WHITE); c.font = font(size=11, bold=True)
    c.alignment = align('center', 'center', False); c.border = BORDER
    c = ws.cell(row=row, column=7); c.value = ''
    c.fill = fill(WHITE); c.border = BORDER
    row += 1

# Rule at the very bottom
ws.row_dimensions[row].height = 12
row += 1
ws.row_dimensions[row].height = 50
c = ws.cell(row=row, column=2); c.value = '⚠ RULE: any "N" in the checklist above = do NOT publish. Fix first, then re-check. Publishing failures is what burned ₹2,733 on the 5 worst creatives.'
c.fill = fill(RED); c.font = font(size=12, bold=True, color=WHITE)
c.alignment = align('center', 'center', True); c.border = BORDER_HDR
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)

ws.freeze_panes = 'A6'
print(f'✓ Sheet 9: Designer Playbook (final row {row}) — built (will save at end)')

# ============================================================
# SHEET 10: MEDIA BUYER RULES
# ============================================================
ws = wb.create_sheet('10. Media Buyer Rules')
ws.sheet_view.showGridLines = False

ws.column_dimensions['A'].width = 2
ws.column_dimensions['B'].width = 4
ws.column_dimensions['C'].width = 24
ws.column_dimensions['D'].width = 22
ws.column_dimensions['E'].width = 22
ws.column_dimensions['F'].width = 50
ws.column_dimensions['G'].width = 24

title(ws, 'B2', 'Media Buyer Decision Rules')
subtitle(ws, 'B3', 'Daily-checked rules for the media buyer / agency. Every rule has a metric trigger and a clear action. No judgment calls.')
subtitle(ws, 'B4', 'These rules are how you avoid the 0.83% CR day that happened on 18 May (₹1,557 spend, 1 sale). Discipline beats new creative.')

# Daily allocation rule
ws.row_dimensions[6].height = 12
section(ws, 'B7', '📅 Rule #1 — Daily allocation discipline')

ws.row_dimensions[9].height = 110
c = ws.cell(row=9, column=2); c.value = '60%'
c.fill = fill(GOLD); c.font = font(size=32, bold=True, color=WHITE)
c.alignment = align('center', 'center', False); c.border = BORDER

c = ws.cell(row=9, column=3); c.value = (
    'AT LEAST 60% of daily ad spend MUST go to the 5 Mixpanel-confirmed winners:\n\n'
    '★ Skill-A72   ★ Skill-V8   ★ Skill-A71   ★ Skill-A63   ★ Self-A30'
)
c.fill = fill(GREEN_LT); c.font = font(size=11, bold=True, color=NAVY)
c.alignment = align('left', 'center', True); c.border = BORDER
ws.merge_cells(start_row=9, start_column=3, end_row=9, end_column=6)

c = ws.cell(row=9, column=7); c.value = (
    'Why: same creative pool gave 0.83% CR on 18 May (when winners were starved) and 5.85% CR on 21 May (when winners got the spend).'
)
c.fill = fill(WHITE); c.font = font(size=10, color='333333', italic=True)
c.alignment = align('left', 'center', True); c.border = BORDER

# New creative launch rules
ws.row_dimensions[11].height = 12
section(ws, 'B12', '🚀 Rule #2 — New creative launch decision table')
subtitle(ws, 'B13', 'Check the new creative\'s metrics at 24 hours, 48 hours, and ₹300 spend. Use this table to decide kill / hold / scale.')

ws.row_dimensions[15].height = 38
hdr = ['#', 'Trigger', 'CPM', 'CTR / Sales', 'What it means', 'Action']
for i, h in enumerate(hdr):
    c = ws.cell(row=15, column=2+i); c.value = h
    c.fill = fill(NAVY); c.font = font(size=10, bold=True, color=WHITE)
    c.alignment = align('center', 'center', True); c.border = BORDER_HDR

launch_rules = [
    ('At 24h after launch', 'CPM < ₹200', 'CTR ≥ 4%, 0 sales',
     'Looks like clickbait or news-style — Meta is serving cheap to wrong audience',
     'PAUSE same day. Audit creative-to-landing-page match.', RED_LT),
    ('At 24h after launch', 'CPM ₹200–280', 'CTR ≥ 3%, ≥ 1 sale',
     'Healthy launch. Auction is normal, hook works, conversion landing.',
     'HOLD at current budget. Re-check at 48h.', AMBER_LT),
    ('At 24h after launch', 'CPM > ₹300', 'CTR < 2.5%',
     'Meta finds it irrelevant AND viewers aren\'t clicking. Bad creative.',
     'PAUSE same day. No salvage.', RED_LT),
    ('At 24h after launch', 'CPM ₹200–250', 'CTR ≥ 3.5%, ≥ 2 sales',
     'Strong opening. All three signals positive.',
     'INCREASE budget 50% same day.', GREEN_LT),
    ('At ₹300 spent', 'Any CPM', '0 sales',
     'No matter how good CPM/CTR look — without sales it doesn\'t pay back.',
     'KILL. Do not throw good money after bad.', RED_LT),
    ('At ₹500 spent', 'Any CPM', '1 sale, CPP > ₹500',
     'Marginal — not yet profitable at the unit level.',
     'CAP at current budget. Don\'t scale until CPP < ₹400.', AMBER_LT),
    ('At ₹500 spent', 'Any CPM', '3+ sales, CR ≥ 4%',
     'Hit the winner signature. Reproduce conditions.',
     'SCALE 2× next day, KEEP same ad ID (no rename).', GREEN_LT),
    ('At 48h+', 'CPM rising > 20% week-over-week', 'CTR also dropping',
     'Audience fatigue setting in. Frequency too high.',
     'CAP budget. Refresh with a sibling creative.', AMBER_LT),
]

for i, (trigger, cpm, ctr_sales, meaning, action, bg) in enumerate(launch_rules):
    r = 16 + i
    ws.row_dimensions[r].height = 50
    c = ws.cell(row=r, column=2); c.value = i+1
    c.fill = fill(NAVY); c.font = font(size=11, bold=True, color=WHITE)
    c.alignment = align('center', 'center', False); c.border = BORDER

    cells = [trigger, cpm, ctr_sales, meaning, action]
    for j, v in enumerate(cells):
        c = ws.cell(row=r, column=3+j); c.value = v
        c.fill = fill(bg)
        if j == 4:
            c.font = font(size=10, bold=True, color=NAVY)
        else:
            c.font = font(size=10, color='333333')
        c.alignment = align('left' if j in (3,4) else 'center', 'center', True)
        c.border = BORDER

# CPM interpretation block
ws.row_dimensions[25].height = 12
section(ws, 'B26', '⚠ Rule #3 — How to read CPM correctly (where the money traps are)')

cpm_rules = [
    ('Very low CPM (< ₹180) + Very high CTR (> 4%) + Zero sales',
     'CLICKBAIT TRAP. Creative is being served to a content-curious audience, not buyers. Meta picks the cheapest reach but it\'s the wrong reach.',
     'KILL immediately. Examples: Skill-A74 (₹135 CPM, 5.1% CTR, 1 sale on 113 visits), Skill-A82 (₹128 CPM, 4.3% CTR, 0 sales).',
     RED_LT),
    ('Low CPM (₹180–220) at high spend (>₹5k)',
     'EARNED LOW CPM. Meta has learned the audience and is delivering efficiently. This is the actual goal state.',
     'KEEP RUNNING. Reuse same ad ID. Examples: Sales Ad - A/B (₹125 CPM at ₹1L+ spend), Still - HM2 (₹133 CPM at ₹12k).',
     GREEN_LT),
    ('Medium CPM (₹220–280) at low spend (<₹3k)',
     'NORMAL TESTING ZONE. Meta hasn\'t finished optimizing yet. Don\'t over-read these numbers.',
     'WAIT until ₹500 spend before judging. Then evaluate on sales, not CPM.',
     AMBER_LT),
    ('High CPM (> ₹300) and good conversion',
     'TIGHT AUDIENCE PAYING OFF. Expensive impressions but they convert. Worth it if CPP stays low.',
     'OK TO RUN — but watch CPP. Examples: Self-A30 (₹396 CPM, 6% CR, ₹288/sale). If CPP creeps above ₹400, pause.',
     GREEN_LT),
    ('High CPM (> ₹300) and zero conversion',
     'WORST CASE. Expensive impressions AND no return.',
     'KILL same day. No iteration. Examples: General-C1 (₹397 CPM, 0 sales, ₹551 spent).',
     RED_LT),
]

ws.row_dimensions[28].height = 38
# Write values first, then merge afterwards
ws.cell(row=28, column=2).value = '#'
ws.cell(row=28, column=3).value = 'CPM + performance signature'
ws.cell(row=28, column=5).value = 'What it really means'
ws.cell(row=28, column=6).value = 'Action + example'
for col_idx in [2, 3, 5, 6]:
    c = ws.cell(row=28, column=col_idx)
    c.fill = fill(NAVY); c.font = font(size=10, bold=True, color=WHITE)
    c.alignment = align('center', 'center', True); c.border = BORDER_HDR
ws.merge_cells(start_row=28, start_column=3, end_row=28, end_column=4)
ws.merge_cells(start_row=28, start_column=6, end_row=28, end_column=7)

for i, (sig, mean, action, bg) in enumerate(cpm_rules):
    r = 29 + i
    ws.row_dimensions[r].height = 78
    c = ws.cell(row=r, column=2); c.value = i+1
    c.fill = fill(NAVY); c.font = font(size=11, bold=True, color=WHITE)
    c.alignment = align('center', 'center', False); c.border = BORDER

    c = ws.cell(row=r, column=3); c.value = sig
    c.fill = fill(bg); c.font = font(size=10, bold=True, color=NAVY)
    c.alignment = align('left', 'center', True); c.border = BORDER
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)

    c = ws.cell(row=r, column=5); c.value = mean
    c.fill = fill(WHITE); c.font = font(size=10, color='333333')
    c.alignment = align('left', 'center', True); c.border = BORDER

    c = ws.cell(row=r, column=6); c.value = action
    c.fill = fill(bg); c.font = font(size=10, bold=True, color='333333')
    c.alignment = align('left', 'center', True); c.border = BORDER
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)

# Fatigue rules
ws.row_dimensions[35].height = 12
section(ws, 'B36', '😴 Rule #4 — Fatigue / refresh triggers (when to retire a winner)')

ws.row_dimensions[38].height = 38
# Write values first, then merge
ws.cell(row=38, column=2).value = 'Signal'
ws.cell(row=38, column=4).value = 'What to watch'
ws.cell(row=38, column=5).value = 'Threshold'
ws.cell(row=38, column=6).value = 'Action'
for col_idx in [2, 4, 5, 6]:
    c = ws.cell(row=38, column=col_idx)
    c.fill = fill(NAVY); c.font = font(size=10, bold=True, color=WHITE)
    c.alignment = align('center', 'center', True); c.border = BORDER_HDR
ws.merge_cells(start_row=38, start_column=2, end_row=38, end_column=3)
ws.merge_cells(start_row=38, start_column=6, end_row=38, end_column=7)

fatigue_rules = [
    ('Frequency rising', 'Same person sees the ad 4+ times', 'Frequency > 4 on the ad set', 'Refresh with a sibling creative same week'),
    ('CPM creeping up', 'Week-over-week increase', 'CPM up > 20% WoW with no audience change', 'Pause and refresh; audience cooling'),
    ('CTR dropping', 'Week-over-week decrease', 'CTR down > 20% WoW', 'Refresh creative with a new headline + new image'),
    ('CPP rising', 'Week-over-week increase', 'CPP up > 30% WoW', 'Cap budget and prepare a new creative variant'),
    ('Running 14+ days at scale', 'Calendar age', '14 days at > ₹500/day daily spend', 'Have a refresh ready to launch on day 15'),
]

for i, (sig, watch, thresh, action) in enumerate(fatigue_rules):
    r = 39 + i
    ws.row_dimensions[r].height = 36
    c = ws.cell(row=r, column=2); c.value = sig
    c.fill = fill(AMBER_LT); c.font = font(size=10, bold=True, color=NAVY)
    c.alignment = align('left', 'center', True); c.border = BORDER
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)

    c = ws.cell(row=r, column=4); c.value = watch
    c.fill = fill(WHITE); c.font = font(size=10, color='333333')
    c.alignment = align('left', 'center', True); c.border = BORDER

    c = ws.cell(row=r, column=5); c.value = thresh
    c.fill = fill(AMBER_LT); c.font = font(size=10, bold=True, color=NAVY)
    c.alignment = align('center', 'center', True); c.border = BORDER

    c = ws.cell(row=r, column=6); c.value = action
    c.fill = fill(WHITE); c.font = font(size=10, color='333333')
    c.alignment = align('left', 'center', True); c.border = BORDER
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)

# This-week action list
ws.row_dimensions[45].height = 12
section(ws, 'B46', '🎯 Rule #5 — This week\'s action list (apply Rules 1-4 to live ads)')

ws.row_dimensions[48].height = 38
# Body uses: col2 act, col3 ad, col4 state, col5-6 move, col7 target
ws.cell(row=48, column=2).value = 'Action'
ws.cell(row=48, column=3).value = 'Ad'
ws.cell(row=48, column=4).value = 'Current state'
ws.cell(row=48, column=5).value = 'Move'
ws.cell(row=48, column=7).value = 'Target'
for col_idx in [2, 3, 4, 5, 7]:
    c = ws.cell(row=48, column=col_idx)
    c.fill = fill(NAVY); c.font = font(size=10, bold=True, color=WHITE)
    c.alignment = align('center', 'center', True); c.border = BORDER_HDR
ws.merge_cells(start_row=48, start_column=5, end_row=48, end_column=6)

this_week = [
    ('SCALE 10×', 'Skill-A72', '₹188 CPM · 7.9% CR · ₹151/sale · ₹1,213 lifetime spend (under-funded)',
     'Push daily budget from ~₹40 to ~₹400 over 5 days. Same ad ID — NO rename.', 'CR ≥ 6%, CPP ≤ ₹250', GREEN_LT),
    ('SCALE 2×', 'Skill-V8', '₹184 CPM · 6.7% CR · ₹156/sale',
     'Increase budget 100% in Testing ad set. Don\'t move to a Winner duplicate.', 'CR ≥ 5%, CPP ≤ ₹250', GREEN_LT),
    ('PROTECT', 'Skill-A71, Skill-A63, Self-A30', 'All converting 5-6% CR at ₹200-300 CPP',
     'Maintain current budgets. Allocate at least ₹200/day to each.', 'CR stays ≥ 5%', GREEN_LT),
    ('KILL TODAY', 'Skill-A82, Skill-A21, Skill-A80, Skill-A74, General-C1',
     '₹2,733 spent combined · 1 sale total · all hit one or more KILL triggers',
     'Pause permanently. Do not iterate or rename.', 'Stop the bleed', RED_LT),
    ('CAP', 'Skill-A68, Skill-A78, Skill-A55', 'Bollywood / tabloid style — mixed results',
     'Cap at ₹500/week each. Pause if no sale in 7 days at cap.', '≥ 1 sale per ₹500', AMBER_LT),
    ('STOP CREATING', '"–Winner" rename duplicates', 'All –Winner duplicates underperform originals',
     'Promote winners by raising the ORIGINAL ad ID\'s budget. No re-uploads.', '0 new –Winner ads', RED_LT),
    ('INSTRUMENT', 'Interest ad, Non Interest ad, 1k Interest ad, 1k Non Interest ad', '₹6,915 spent · NO landing-page tracking · funnel invisible',
     'Add LP instrumentation OR pause. Cannot judge what we cannot measure.', 'Funnel visible OR ads off', GREY_LT),
]

for i, (act, ad, state, move, target, bg) in enumerate(this_week):
    r = 49 + i
    ws.row_dimensions[r].height = 56
    c = ws.cell(row=r, column=2); c.value = act
    c.fill = fill(bg); c.font = font(size=11, bold=True, color=NAVY)
    c.alignment = align('center', 'center', True); c.border = BORDER

    c = ws.cell(row=r, column=3); c.value = ad
    c.fill = fill(bg); c.font = font(size=10, bold=True, color=NAVY)
    c.alignment = align('left', 'center', True); c.border = BORDER

    c = ws.cell(row=r, column=4); c.value = state
    c.fill = fill(WHITE); c.font = font(size=10, color='333333')
    c.alignment = align('left', 'center', True); c.border = BORDER

    c = ws.cell(row=r, column=5); c.value = move
    c.fill = fill(bg); c.font = font(size=10, bold=True, color=NAVY)
    c.alignment = align('left', 'center', True); c.border = BORDER
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)

    c = ws.cell(row=r, column=7); c.value = target
    c.fill = fill(WHITE); c.font = font(size=10, color='333333', italic=True)
    c.alignment = align('center', 'center', True); c.border = BORDER

# Sanity check at the bottom
ws.row_dimensions[57].height = 12
ws.row_dimensions[58].height = 56
c = ws.cell(row=58, column=2); c.value = (
    '⚠ DAILY SANITY CHECK before increasing any budget:\n'
    '   1) Is at least 60% of daily spend on the 5 Winners?   2) Are all KILL-list ads actually paused?   3) Is every new ad past its 24h trigger threshold?\n'
    'If any answer is NO — fix that before increasing any budget. Discipline beats creativity in paid media.'
)
c.fill = fill(NAVY); c.font = font(size=11, bold=True, color=WHITE)
c.alignment = align('left', 'center', True); c.border = BORDER_HDR
ws.merge_cells(start_row=58, start_column=2, end_row=58, end_column=7)

ws.freeze_panes = 'A6'
print(f'✓ Sheet 10: Media Buyer Rules — built (will save at end)')

# ============================================================
# Update Executive Summary navigation to include new sheets
# ============================================================
ws = wb['1. Executive Summary']
# Find the workbook contents section and add 2 new rows
# (We'll just append two rows at row 42 and 43)
new_entries = [
    (42, '9. Designer Playbook', 'Hand-to-designer brief · visual side-by-side · DO/DON\'T checklists · pre-launch QA gate'),
    (43, '10. Media Buyer Rules', 'Decision triggers · daily allocation · launch / kill / scale rules · this week\'s action list'),
]
for row, name, desc in new_entries:
    ws.row_dimensions[row].height = 22
    c = ws.cell(row=row, column=2); c.value = '→'
    c.font = font(size=14, bold=True, color=GOLD)
    c.alignment = align('center', 'center', False)

    c = ws.cell(row=row, column=3); c.value = name
    c.font = font(size=11, bold=True, color=NAVY)
    c.alignment = align('left', 'center', False)
    c.fill = fill(GREY_LT); c.border = BORDER

    c = ws.cell(row=row, column=4); c.value = desc
    c.font = font(size=10, color='333333')
    c.alignment = align('left', 'center', False)
    c.fill = fill(WHITE); c.border = BORDER

wb.save(XLSX)
print(f'\n✅ All sheets saved to {XLSX}')
print(f'   Sheets in workbook: {wb.sheetnames}')
print(f'   File size: {os.path.getsize(XLSX)/1024:.0f} KB')
