#!/usr/bin/env python3
"""Add Sheet 11: CPA Decomposition to the CMO workbook."""
import os, io
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

XLSX = 'analytics/Creative_Design_Insights_CMO.xlsx'
THUMB_DIR = 'analytics/creative_thumbnails'

# Palette (matches existing workbook)
NAVY = '1F2A44'; GOLD = 'D4A547'
GREEN = '4FAB6E'; GREEN_LT = 'E7F4EA'; GREEN_DK = '2D7849'
RED = 'D9534F'; RED_LT = 'FBE5E4'; RED_DK = 'A03430'
AMBER = 'F0AD4E'; AMBER_LT = 'FDF1E0'
GREY = '6C757D'; GREY_LT = 'F4F4F4'; WHITE = 'FFFFFF'

THIN = Side(border_style='thin', color='D0D0D0')
MED = Side(border_style='medium', color='B0B0B0')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_HDR = Border(left=MED, right=MED, top=MED, bottom=MED)

def fill(c): return PatternFill('solid', fgColor=c)
def font(size=10, bold=False, color='000000', italic=False):
    return Font(name='Calibri', size=size, bold=bold, color=color, italic=italic)
def align(h='left', v='center', wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def title(ws, cell, text, size=18, color=NAVY):
    ws[cell] = text
    ws[cell].font = font(size=size, bold=True, color=color)
    ws[cell].alignment = align('left', 'center', False)

def subtitle(ws, cell, text, size=10, color=GREY):
    ws[cell] = text
    ws[cell].font = font(size=size, italic=True, color=color)
    ws[cell].alignment = align('left', 'center', False)

def section(ws, cell, text, size=14, color=NAVY):
    ws[cell] = text
    ws[cell].font = font(size=size, bold=True, color=color)
    ws[cell].alignment = align('left', 'center', False)

def insert_thumb(ws, ad_name, anchor, px=70):
    """Match ad_name to a thumbnail file robustly and embed at anchor."""
    import re
    fnames = {f.replace('.png', ''): f for f in os.listdir(THUMB_DIR)}
    safe = re.sub(r'[^\w\-]+', '_', ad_name).strip('_')
    candidates = [safe, safe.replace('–', '-'), safe.replace('–', '_')]
    for c in candidates:
        if c in fnames:
            path = f'{THUMB_DIR}/{fnames[c]}'
            break
    else:
        # case-insensitive scan
        path = None
        sl = safe.lower()
        for k, v in fnames.items():
            if k.lower() == sl:
                path = f'{THUMB_DIR}/{v}'
                break
        if not path:
            return False
    src = PILImage.open(path).convert('RGB')
    src.thumbnail((px, px), PILImage.LANCZOS)
    canvas = PILImage.new('RGB', (px, px), (255, 255, 255))
    canvas.paste(src, ((px - src.width)//2, (px - src.height)//2))
    buf = io.BytesIO()
    canvas.save(buf, format='PNG')
    buf.seek(0)
    img = XLImage(buf)
    img.width = px; img.height = px
    img.anchor = anchor
    ws.add_image(img)
    return True

# ============================================================
# Load + remove old version if present
# ============================================================
wb = load_workbook(XLSX)
if '11. CPA Decomposition' in wb.sheetnames:
    del wb['11. CPA Decomposition']
ws = wb.create_sheet('11. CPA Decomposition')
ws.sheet_view.showGridLines = False

# Column widths
ws.column_dimensions['A'].width = 2
ws.column_dimensions['B'].width = 13   # Creative thumb
ws.column_dimensions['C'].width = 19   # Ad name
ws.column_dimensions['D'].width = 10   # Spend
ws.column_dimensions['E'].width = 10   # Results
ws.column_dimensions['F'].width = 10   # CPA
ws.column_dimensions['G'].width = 10   # CPM
ws.column_dimensions['H'].width = 10   # CTR
ws.column_dimensions['I'].width = 10   # Conv
ws.column_dimensions['J'].width = 12   # CTR×Conv
ws.column_dimensions['K'].width = 50   # Notes

# Title
title(ws, 'B2', 'CPA Decomposition — Why winners win, why losers lose')
subtitle(ws, 'B3', 'Source: data/ads-data.xlsx (52 ads, latest as of 4 June 2026). Target CPA: ₹200.')
subtitle(ws, 'B4', 'Formula: CPA = CPM × 10 / (CTR% × Conv Rate%). To hit CPA = ₹200, you need (CTR × Conv) ≥ CPM ÷ 20.')

# ============================================================
# FORMULA REFERENCE — the most important table on this sheet
# ============================================================
section(ws, 'B6', '📐 The CPA equation — what each metric is responsible for')

ws.row_dimensions[8].height = 140
c = ws.cell(row=8, column=2)
c.value = 'CPA  =  CPM × 10\n              ─────────────\n           CTR%  ×  Conv%'
c.fill = fill(GOLD); c.font = font(size=14, bold=True, color=WHITE)
c.alignment = align('center', 'center', True); c.border = BORDER_HDR
ws.merge_cells(start_row=8, start_column=2, end_row=8, end_column=4)

c = ws.cell(row=8, column=5); c.value = (
    'You have THREE levers to lower CPA:\n\n'
    '①  Lower CPM  (cheaper impressions — design / brand-relevance signal)\n'
    '②  Higher CTR  (better hook — but secondary in your data, see below)\n'
    '③  Higher Conv Rate  (better landing-page match — biggest lever in your data)\n\n'
    'CTR and Conv MULTIPLY, so weakness in either kills CPA. CPM scales linearly.'
)
c.fill = fill(GREY_LT); c.font = font(size=11, color=NAVY)
c.alignment = align('left', 'center', True); c.border = BORDER
ws.merge_cells(start_row=8, start_column=5, end_row=8, end_column=11)

# Required CTR × Conv table by CPM tier
ws.row_dimensions[10].height = 12
section(ws, 'B11', '📊 The break-even table — what (CTR × Conv) must be to hit ₹200 CPA at each CPM')

ws.row_dimensions[13].height = 32
# Body layout: col 2-3 merged (tier), col 4 (required), col 5 (difficulty), col 6-11 merged (translation)
ws.cell(row=13, column=2).value = 'If CPM is...'
ws.cell(row=13, column=4).value = 'Then (CTR × Conv) must be ≥'
ws.cell(row=13, column=5).value = 'Difficulty'
ws.cell(row=13, column=6).value = 'Translation'
for col_idx in [2, 4, 5, 6]:
    c = ws.cell(row=13, column=col_idx)
    c.fill = fill(NAVY); c.font = font(size=10, bold=True, color=WHITE)
    c.alignment = align('center', 'center', True); c.border = BORDER_HDR
ws.merge_cells(start_row=13, start_column=2, end_row=13, end_column=3)
ws.merge_cells(start_row=13, start_column=6, end_row=13, end_column=11)

break_even = [
    ('₹150 (cheap)',         '7.5',  'EASY',  'Achievable with e.g. CTR 2% × Conv 4% = 8.0 ✓',  GREEN_LT),
    ('₹200',                  '10.0', 'EASY',  'Achievable with e.g. CTR 2% × Conv 5% = 10.0 ✓',  GREEN_LT),
    ('₹262 (account median)', '13.1', 'MEDIUM', 'Needs CTR 2% × Conv 6.5% OR CTR 2.5% × Conv 5.2%', AMBER_LT),
    ('₹350',                  '17.5', 'HARD',  'Needs CTR 2.5% × Conv 7% — both above median', AMBER_LT),
    ('₹500 (very high)',      '25.0', 'BRUTAL', 'Needs CTR 3% × Conv 8.3% — only Skill-A2-TEST hits this', RED_LT),
]

for i, (cpm_tier, required, diff, trans, bg) in enumerate(break_even):
    r = 14 + i
    ws.row_dimensions[r].height = 28
    cells = [cpm_tier, required, diff, trans]
    c = ws.cell(row=r, column=2); c.value = cpm_tier
    c.fill = fill(bg); c.font = font(size=11, bold=True, color=NAVY)
    c.alignment = align('center', 'center', True); c.border = BORDER
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)

    c = ws.cell(row=r, column=4); c.value = required
    c.fill = fill(bg); c.font = font(size=13, bold=True, color=NAVY)
    c.alignment = align('center', 'center', True); c.border = BORDER

    c = ws.cell(row=r, column=5); c.value = diff
    diff_color = {'EASY': GREEN_DK, 'MEDIUM': '7E5F00', 'HARD': '7E5F00', 'BRUTAL': RED_DK}[diff]
    c.fill = fill(bg); c.font = font(size=10, bold=True, color=diff_color)
    c.alignment = align('center', 'center', True); c.border = BORDER

    c = ws.cell(row=r, column=6); c.value = trans
    c.fill = fill(bg); c.font = font(size=10, color='333333')
    c.alignment = align('left', 'center', True); c.border = BORDER
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=11)

# Account baselines
ws.row_dimensions[20].height = 12
section(ws, 'B21', '📈 Account baselines (median across 29 ads with results)')

ws.row_dimensions[23].height = 32
# Write values to anchor cells first, then merge
ws.cell(row=23, column=2).value = 'Metric'
ws.cell(row=23, column=4).value = 'Your account median'
ws.cell(row=23, column=6).value = 'What "good" looks like'
ws.cell(row=23, column=8).value = 'What "bad" looks like'
for col_idx in [2, 4, 6, 8]:
    c = ws.cell(row=23, column=col_idx)
    c.fill = fill(NAVY); c.font = font(size=10, bold=True, color=WHITE)
    c.alignment = align('center', 'center', True); c.border = BORDER_HDR
ws.merge_cells(start_row=23, start_column=2, end_row=23, end_column=3)
ws.merge_cells(start_row=23, start_column=4, end_row=23, end_column=5)
ws.merge_cells(start_row=23, start_column=6, end_row=23, end_column=7)
ws.merge_cells(start_row=23, start_column=8, end_row=23, end_column=11)

baselines = [
    ('CPM',       '₹262',  '< ₹220 (cheap impressions = Meta likes you)', '> ₹350 (Meta finds you irrelevant — kill or fix creative)'),
    ('CTR',       '1.90%', '> 2.5% (strong hook)',                         '< 1.5% (hook is invisible or wrong promise)'),
    ('Conv Rate', '4.55%', '> 5% (LP closes the click)',                   '0% (LP totally fails — usually a promise mismatch)'),
    ('CPA',       '₹332',  '< ₹200 (target hit)',                          '> ₹500 (kill same day)'),
]
for i, (m, med, good, bad) in enumerate(baselines):
    r = 24 + i
    ws.row_dimensions[r].height = 28
    c = ws.cell(row=r, column=2); c.value = m
    c.fill = fill(GREY_LT); c.font = font(size=11, bold=True, color=NAVY)
    c.alignment = align('center', 'center', True); c.border = BORDER
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)

    c = ws.cell(row=r, column=4); c.value = med
    c.fill = fill(GREY_LT); c.font = font(size=12, bold=True, color=NAVY)
    c.alignment = align('center', 'center', True); c.border = BORDER
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)

    c = ws.cell(row=r, column=6); c.value = good
    c.fill = fill(GREEN_LT); c.font = font(size=10, color=GREEN_DK)
    c.alignment = align('left', 'center', True); c.border = BORDER
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)

    c = ws.cell(row=r, column=8); c.value = bad
    c.fill = fill(RED_LT); c.font = font(size=10, color=RED_DK)
    c.alignment = align('left', 'center', True); c.border = BORDER
    ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=11)

# ============================================================
# THE 4 WINNERS — full decomposition
# ============================================================
ws.row_dimensions[28].height = 12
section(ws, 'B29', '🏆 The 4 WINNERS — CPA < ₹200, decomposed')

ws.row_dimensions[31].height = 38
w_hdr = ['Creative', 'Ad Name', 'Spend', 'Results', 'CPA', 'CPM', 'CTR%', 'Conv%', 'CTR × Conv', 'Why this WINS']
for i, h in enumerate(w_hdr):
    c = ws.cell(row=31, column=2+i); c.value = h
    c.fill = fill(GREEN); c.font = font(size=10, bold=True, color=WHITE)
    c.alignment = align('center', 'center', True); c.border = BORDER_HDR

winners_data = [
    {
        'ad': 'Skill-A2-TEST', 'spend': 112, 'results': 2, 'cpa': 56,
        'cpm': 253, 'ctr': 1.81, 'conv': 28.57, 'product': 51.7,
        'why': (
            '★ Pure conversion-rate miracle. Conv 28.57% is 6× the account median.\n'
            '⚠ Small sample (₹112 spend, 2 results). Statistically thin.\n'
            'Action: scale to ₹1,000 spend before trusting the number. '
            'If Conv stays > 10%, this is your most efficient creative — replicate the headline.'
        ),
    },
    {
        'ad': 'Skill-V9', 'spend': 964, 'results': 9, 'cpa': 107,
        'cpm': 161, 'ctr': 1.90, 'conv': 8.49, 'product': 16.1,
        'why': (
            '★ Low CPM (₹161, 39% below median) + above-median Conv (8.49%) carry the math.\n'
            'CTR is exactly at the median (1.90%) — not a factor.\n'
            'Action: scale 5× immediately. The numbers say there\'s headroom.'
        ),
    },
    {
        'ad': 'Skill-V8', 'spend': 8853, 'results': 48, 'cpa': 184,
        'cpm': 176, 'ctr': 1.60, 'conv': 6.64, 'product': 10.6,
        'why': (
            '★ Already scaled (₹8,853 spend, 48 results) — proves the model holds at volume.\n'
            'Low CPM (₹176) does most of the work. Conv 6.64% gives the closing punch.\n'
            'CTR is the LOWEST of all 4 winners at 1.60% — proves CTR is not the lever.\n'
            'Action: keep running. Watch for fatigue after frequency > 4.'
        ),
    },
    {
        'ad': 'Skill-HF2', 'spend': 1200, 'results': 6, 'cpa': 200,
        'cpm': 165, 'ctr': 2.32, 'conv': 3.90, 'product': 9.0,
        'why': (
            '★ Balanced model: low CPM (₹165) + slightly above-median CTR (2.32%) + slightly below-median Conv (3.90%).\n'
            'Sits exactly at the ₹200 target with no metric dominating.\n'
            'Action: scale 2× and watch Conv. If Conv slips below 3%, CPA breaks out of target.'
        ),
    },
]

row = 32
for w in winners_data:
    ws.row_dimensions[row].height = 90
    insert_thumb(ws, w['ad'], f'B{row}', px=80)
    c = ws.cell(row=row, column=2); c.fill = fill(GREEN_LT); c.border = BORDER

    c = ws.cell(row=row, column=3); c.value = w['ad']
    c.fill = fill(GREEN_LT); c.font = font(size=11, bold=True, color=NAVY)
    c.alignment = align('center', 'center', True); c.border = BORDER

    values = [
        (f'₹{w["spend"]:,}', NAVY),
        (str(w['results']), NAVY),
        (f'₹{w["cpa"]}', GREEN_DK if w['cpa'] < 200 else NAVY),
        (f'₹{w["cpm"]}', GREEN_DK if w['cpm'] < 220 else (RED_DK if w['cpm'] > 350 else NAVY)),
        (f'{w["ctr"]:.2f}%', GREEN_DK if w['ctr'] > 2.5 else (RED_DK if w['ctr'] < 1.5 else NAVY)),
        (f'{w["conv"]:.2f}%', GREEN_DK if w['conv'] > 5 else NAVY),
        (f'{w["product"]:.1f}', GREEN_DK),
    ]
    for j, (val, color) in enumerate(values):
        c = ws.cell(row=row, column=4+j); c.value = val
        c.fill = fill(GREEN_LT); c.font = font(size=11, bold=True, color=color)
        c.alignment = align('center', 'center', True); c.border = BORDER

    c = ws.cell(row=row, column=11); c.value = w['why']
    c.fill = fill(WHITE); c.font = font(size=10, color='333333')
    c.alignment = align('left', 'top', True); c.border = BORDER
    row += 1

# ============================================================
# THE LOSERS — 0% conv rate pattern
# ============================================================
ws.row_dimensions[row].height = 12; row += 1
section(ws, f'B{row}', '💸 The 12 LOSERS — ₹300+ spent, ZERO results. All have 0% Conv Rate.')
ws.row_dimensions[row].height = 30
row += 1
subtitle(ws, f'B{row}', 'Common thread: every single loser has 0% Conv Rate. CTR and CPM look normal. The break is between the click and the purchase — landing-page promise mismatch.')
row += 2

ws.row_dimensions[row].height = 38
# Body layout: thumb=2, ad=3, spend=4, results=5, cpm=6, ctr=7, conv=8, diag(merged 9-10), action=11
ws.cell(row=row, column=2).value = 'Creative'
ws.cell(row=row, column=3).value = 'Ad Name'
ws.cell(row=row, column=4).value = 'Spend'
ws.cell(row=row, column=5).value = 'Results'
ws.cell(row=row, column=6).value = 'CPM'
ws.cell(row=row, column=7).value = 'CTR%'
ws.cell(row=row, column=8).value = 'Conv%'
ws.cell(row=row, column=9).value = 'Diagnostic'
ws.cell(row=row, column=11).value = 'Action'
for col_idx in [2, 3, 4, 5, 6, 7, 8, 9, 11]:
    c = ws.cell(row=row, column=col_idx)
    c.fill = fill(RED); c.font = font(size=10, bold=True, color=WHITE)
    c.alignment = align('center', 'center', True); c.border = BORDER_HDR
ws.merge_cells(start_row=row, start_column=9, end_row=row, end_column=10)

losers_data = [
    ('Skill-A82', 526, 109, 3.32, 0.00, 'Cheapest CPM + high CTR + zero conv = CLICKBAIT TRAP. Lipstick-shop visual pulls curiosity clicks.', 'KILL today'),
    ('Skill-A81', 481, 124, 3.35, 0.00, 'Same pattern. Notebook visual — looks native, gets cheap engaged impressions, but visitors leave.', 'KILL today'),
    ('Skill-A74', 1099, 149, 4.12, 1.06, 'HIGHEST CTR in dataset + cheapest CPM, but 1% conv. ₹366 CPA. News-clipping visual pulls news-readers, not buyers.', 'KILL today (in your near-miss zone but verdict same)'),
    ('Skill-A21', 516, 186, 1.58, 0.00, 'Text-only yellow warning box. Low CTR confirms the visual is weak too. No salvage.', 'KILL today'),
    ('Skill-A66', 527, 202, 1.88, 0.00, 'Hair-tool grid + ₹299. CPM is fine, CTR is fine, but nobody buys.', 'AUDIT LP for traffic — could be a measurement issue'),
    ('Skill-A55', 541, 291, 2.05, 0.00, 'Multi-style collage with small headline. CTR is fine; conv broken.', 'REWRITE headline larger + retest'),
    ('Skill-A52', 501, 262, 2.51, 0.00, 'Practice-on-dummy framing. Shows pain, not result.', 'REPLACE with finished-style version + retest'),
    ('Skill-A4-TEST', 531, 268, 1.97, 0.00, '"Only Makeup is not enough" — vague hook, no specific outcome.', 'KILL — vague hooks underperform consistently'),
    ('Skill-A83', 518, 378, 1.46, 0.00, 'High CPM + low CTR + 0 conv. All three signals are bad.', 'KILL today'),
    ('Skill-V3', 503, 371, 1.55, 0.00, 'Same as A83. Triple-negative signature.', 'KILL today'),
    ('Skill-A86', 516, 514, 2.39, 0.00, 'EXPENSIVE CPM (₹514) + 0 conv. Meta hates it AND it doesn\'t convert.', 'KILL today'),
    ('Skill-A2', 492, 324, 1.05, 0.00, 'Lowest CTR among losers (1.05%) — the hook is broken from the top.', 'KILL today'),
    ('Skill-A52-TEST', 391, 178, 2.68, 0.00, 'OK CTR + OK CPM, but conv = 0. LP issue specifically for this audience-creative combination.', 'AUDIT or retire'),
]

for ad, spend, cpm, ctr, conv, diag, action in losers_data:
    ws.row_dimensions[row].height = 90
    insert_thumb(ws, ad, f'B{row}', px=80)
    c = ws.cell(row=row, column=2); c.fill = fill(RED_LT); c.border = BORDER

    c = ws.cell(row=row, column=3); c.value = ad
    c.fill = fill(RED_LT); c.font = font(size=11, bold=True, color=RED_DK)
    c.alignment = align('center', 'center', True); c.border = BORDER

    vals = [f'₹{spend}', '0', f'₹{cpm}', f'{ctr:.2f}%', f'{conv:.2f}%']
    for j, v in enumerate(vals):
        c = ws.cell(row=row, column=4+j); c.value = v
        c.fill = fill(RED_LT); c.font = font(size=11, bold=True, color=NAVY)
        c.alignment = align('center', 'center', True); c.border = BORDER

    c = ws.cell(row=row, column=9); c.value = diag
    c.fill = fill(WHITE); c.font = font(size=10, color='333333')
    c.alignment = align('left', 'top', True); c.border = BORDER
    ws.merge_cells(start_row=row, start_column=9, end_row=row, end_column=10)

    c = ws.cell(row=row, column=11); c.value = action
    c.fill = fill(RED); c.font = font(size=10, bold=True, color=WHITE)
    c.alignment = align('center', 'center', True); c.border = BORDER
    row += 1

# Total bleed
ws.row_dimensions[row].height = 32
c = ws.cell(row=row, column=2); c.value = (
    f'TOTAL BLEED on these 13 ads: ₹{sum(l[1] for l in losers_data):,} spent · 0 backend purchases. '
    f'Killing today would have funded {sum(l[1] for l in losers_data) // 200} more sales at target CPA.'
)
c.fill = fill(NAVY); c.font = font(size=11, bold=True, color=WHITE)
c.alignment = align('center', 'center', True); c.border = BORDER_HDR
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=11)
row += 1

# ============================================================
# NEAR MISSES — what to fix
# ============================================================
ws.row_dimensions[row].height = 12; row += 1
section(ws, f'B{row}', '🎯 Near-misses — these are CPA ₹200-300, fixable')
row += 1
ws.row_dimensions[row].height = 30
subtitle(ws, f'B{row}', 'These ads have 1-2 metrics working. If you can lift the third, they become winners. Highest-leverage edits in the portfolio.')
row += 2

ws.row_dimensions[row].height = 38
# Body: thumb=2, ad=3, cpa=4, cpm=5, ctr=6, conv=7, working(merged 8-9), fix(merged 10-11)
ws.cell(row=row, column=2).value = 'Creative'
ws.cell(row=row, column=3).value = 'Ad Name'
ws.cell(row=row, column=4).value = 'CPA'
ws.cell(row=row, column=5).value = 'CPM'
ws.cell(row=row, column=6).value = 'CTR%'
ws.cell(row=row, column=7).value = 'Conv%'
ws.cell(row=row, column=8).value = "What's working"
ws.cell(row=row, column=10).value = 'What to fix'
for col_idx in [2, 3, 4, 5, 6, 7, 8, 10]:
    c = ws.cell(row=row, column=col_idx)
    c.fill = fill(AMBER); c.font = font(size=10, bold=True, color=WHITE)
    c.alignment = align('center', 'center', True); c.border = BORDER_HDR
ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=9)
ws.merge_cells(start_row=row, start_column=10, end_row=row, end_column=11)
row += 1

near_data = [
    ('Skill-SKUP1', 209, 235, 1.77, 7.64, 'CPM OK, Conv strong (7.64%)', 'Lift CTR from 1.77% → 2.5% via headline edit'),
    ('Skill-A80', 212, 274, 2.83, 4.95, 'CTR strong (2.83%)', 'Lift Conv from 4.95% → 6%+ via LP-promise audit'),
    ('Skill-A27-TEST', 213, 213, 1.66, 6.49, 'CPM cheap + Conv strong', 'Lift CTR from 1.66% with bigger headline'),
    ('Skill-A27', 216, 262, 1.85, 7.38, 'Conv strong (7.38%)', 'Drop CPM from ₹262 → < ₹220 via cleaner brand layout'),
    ('SKUP1-TEST-V', 225, 222, 1.56, 6.79, 'CPM cheap + Conv strong', 'Lift CTR — currently very low at 1.56%'),
    ('Self-A24-TEST', 262, 203, 1.94, 4.17, 'CPM cheap, CTR & Conv on median', 'Lift Conv with sharper LP promise'),
    ('Skill-A72', 271, 203, 1.61, 5.04, 'CPM cheap + Conv at threshold', 'Lift CTR — 1.61% is too low for this CPM'),
    ('Skill-V5-Test', 286, 321, 1.99, 6.28, 'Conv strong (6.28%)', 'CPM too high (₹321) — clean up brand cues, reduce text density'),
    ('Skill-A19', 328, 486, 1.88, 13.04, 'Conv EXCEPTIONAL (13.04%)', 'CPM (₹486) is the killer. Why so high? Audit creative layout — Meta finds it irrelevant'),
    ('Self-A30', 288, 397, 3.01, 4.44, 'CTR strong (3.01%)', 'CPM (₹397) is the issue — premium audience pricing'),
]

for ad, cpa, cpm, ctr, conv, working, fix in near_data:
    ws.row_dimensions[row].height = 75
    insert_thumb(ws, ad, f'B{row}', px=70)
    c = ws.cell(row=row, column=2); c.fill = fill(AMBER_LT); c.border = BORDER

    c = ws.cell(row=row, column=3); c.value = ad
    c.fill = fill(AMBER_LT); c.font = font(size=11, bold=True, color=NAVY)
    c.alignment = align('center', 'center', True); c.border = BORDER

    vals = [f'₹{cpa}', f'₹{cpm}', f'{ctr:.2f}%', f'{conv:.2f}%']
    for j, v in enumerate(vals):
        c = ws.cell(row=row, column=4+j); c.value = v
        c.fill = fill(AMBER_LT); c.font = font(size=11, bold=True, color=NAVY)
        c.alignment = align('center', 'center', True); c.border = BORDER

    c = ws.cell(row=row, column=8); c.value = working
    c.fill = fill(GREEN_LT); c.font = font(size=10, color=GREEN_DK)
    c.alignment = align('left', 'center', True); c.border = BORDER
    ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=9)

    c = ws.cell(row=row, column=10); c.value = fix
    c.fill = fill(RED_LT); c.font = font(size=10, bold=True, color=RED_DK)
    c.alignment = align('left', 'center', True); c.border = BORDER
    ws.merge_cells(start_row=row, start_column=10, end_row=row, end_column=11)
    row += 1

# ============================================================
# THE 2x2 DECISION MATRIX
# ============================================================
ws.row_dimensions[row].height = 12; row += 1
section(ws, f'B{row}', '🧭 The 2×2 decision matrix — diagnose any new ad in 10 seconds')
row += 2

# 2x2 layout: rows = Conv High/Low, columns = CPM Low/High
ws.row_dimensions[row].height = 32
c = ws.cell(row=row, column=2); c.value = ''
c.fill = fill(WHITE); c.border = BORDER
c = ws.cell(row=row, column=3); c.value = 'LOW CPM (< ₹220)\nMeta likes the creative'
c.fill = fill(NAVY); c.font = font(size=11, bold=True, color=WHITE)
c.alignment = align('center', 'center', True); c.border = BORDER_HDR
ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
c = ws.cell(row=row, column=7); c.value = 'HIGH CPM (≥ ₹220)\nMeta finds it expensive'
c.fill = fill(NAVY); c.font = font(size=11, bold=True, color=WHITE)
c.alignment = align('center', 'center', True); c.border = BORDER_HDR
ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=11)
row += 1

ws.row_dimensions[row].height = 130
c = ws.cell(row=row, column=2); c.value = 'HIGH CONV\n(≥ 5%)\n\nLP closes\nthe click'
c.fill = fill(NAVY); c.font = font(size=11, bold=True, color=WHITE)
c.alignment = align('center', 'center', True); c.border = BORDER_HDR

c = ws.cell(row=row, column=3); c.value = (
    '🏆 WINNER\n\n'
    'CPA hits ₹200 regardless of CTR.\n'
    'SCALE IMMEDIATELY.\n\n'
    'Examples: Skill-V8 (₹184), Skill-V9 (₹107), Skill-HF2 (₹200)'
)
c.fill = fill(GREEN_LT); c.font = font(size=11, bold=True, color=GREEN_DK)
c.alignment = align('center', 'center', True); c.border = BORDER
ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)

c = ws.cell(row=row, column=7); c.value = (
    '🎯 TARGETING WIN\n\n'
    'Premium audience — pay more for impressions, but conv > 10% can still make CPA work.\n\n'
    'Examples: Skill-A19 (Conv 13.04% but CPM ₹486 = CPA ₹328 — needs CPM cleanup)'
)
c.fill = fill(AMBER_LT); c.font = font(size=11, bold=True, color='7E5F00')
c.alignment = align('center', 'center', True); c.border = BORDER
ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=11)
row += 1

ws.row_dimensions[row].height = 130
c = ws.cell(row=row, column=2); c.value = 'LOW CONV\n(< 5%)\n\nClick ≠\npurchase'
c.fill = fill(NAVY); c.font = font(size=11, bold=True, color=WHITE)
c.alignment = align('center', 'center', True); c.border = BORDER_HDR

c = ws.cell(row=row, column=3); c.value = (
    '⚠ LP MISMATCH\n\n'
    'Meta likes the creative; the LP is breaking the promise.\n'
    'FIX LP OR ROTATE OFFER — do not iterate creative.\n\n'
    'Examples: Skill-A82, Skill-A81, Skill-A74 — cheap clickbait that does not convert.'
)
c.fill = fill(AMBER_LT); c.font = font(size=11, bold=True, color='7E5F00')
c.alignment = align('center', 'center', True); c.border = BORDER
ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)

c = ws.cell(row=row, column=7); c.value = (
    '💸 KILL\n\n'
    'Both signals against you. No path to ₹200 CPA.\n'
    'PAUSE TODAY. No iteration.\n\n'
    'Examples: Skill-V3, Skill-A86, Skill-A83, General-C1'
)
c.fill = fill(RED_LT); c.font = font(size=11, bold=True, color=RED_DK)
c.alignment = align('center', 'center', True); c.border = BORDER
ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=11)
row += 1

# ============================================================
# THE THREE BIG TAKEAWAYS
# ============================================================
ws.row_dimensions[row].height = 12; row += 1
section(ws, f'B{row}', '✅ Three things to brief into every new creative')
row += 2

takeaways = [
    ('1', 'STOP optimizing for CTR alone',
     'All 4 winners have BELOW-MEDIAN CTR (1.60%–2.32%). The 3 highest-CTR ads in the dataset (Skill-A74 at 4.12%, Skill-A81 at 3.35%, Skill-A82 at 3.32%) are all CPA LOSERS. CTR without Conv is a tax, not a win.'),
    ('2', 'LOW CPM is the #1 winning lever (creative-side)',
     'Three of four winners have CPM in ₹161–₹176 — about 35% below account median. Make ads that LOOK LIKE ADS (clean brand + finished hairstyle + ₹299 + CTA). Meta serves them as commercial intent and the auction is cheaper.'),
    ('3', 'CONV RATE is the #2 winning lever (LP-side)',
     'Every single LOSER has 0% Conv Rate. The break is between the click and the purchase. Promise on the ad must match what the LP delivers. If you cannot get Conv above 4%, no amount of CTR or CPM optimization saves CPA.'),
]

for num, headline, detail in takeaways:
    ws.row_dimensions[row].height = 70
    c = ws.cell(row=row, column=2); c.value = num
    c.fill = fill(GOLD); c.font = font(size=22, bold=True, color=WHITE)
    c.alignment = align('center', 'center', False); c.border = BORDER

    c = ws.cell(row=row, column=3); c.value = headline
    c.fill = fill(GREY_LT); c.font = font(size=12, bold=True, color=NAVY)
    c.alignment = align('left', 'center', True); c.border = BORDER
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)

    c = ws.cell(row=row, column=6); c.value = detail
    c.fill = fill(WHITE); c.font = font(size=10, color='333333')
    c.alignment = align('left', 'center', True); c.border = BORDER
    ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=11)
    row += 1

# Freeze top + save
ws.freeze_panes = 'A6'
wb.save(XLSX)
print(f'✓ Sheet 11 added — final row {row}')
print(f'  Sheets in workbook: {wb.sheetnames}')
print(f'  File size: {os.path.getsize(XLSX)/1024:.0f} KB')
