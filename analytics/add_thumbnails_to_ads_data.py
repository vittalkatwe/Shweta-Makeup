#!/usr/bin/env python3
"""Insert a 'Creative' column with embedded thumbnails at the start of ads-data.xlsx."""
import os, re, io
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

XLSX = 'analytics/ads-data.xlsx'
THUMB_DIR = 'analytics/creative_thumbnails'
THUMB_PX = 80  # square pixel size for embedded thumbnails

# Style helpers
THIN = Side(border_style='thin', color='D0D0D0')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Build thumbnail filename lookup (case-insensitive, robust to dashes/spaces)
thumbnails_on_disk = {f.replace('.png', '') for f in os.listdir(THUMB_DIR)}

def find_thumbnail(ad_name):
    """Find matching thumbnail filename for an ad name."""
    if not ad_name:
        return None
    # Strategy 1: exact match after safe-naming
    safe = re.sub(r'[^\w\-]+', '_', ad_name).strip('_')
    if safe in thumbnails_on_disk:
        return f'{THUMB_DIR}/{safe}.png'
    # Strategy 2: normalize en-dashes to hyphens, hyphens to underscores
    for variant in [ad_name.replace('–', '-'), ad_name.replace('-', '_'), ad_name.replace('–', '_')]:
        v = re.sub(r'[^\w\-]+', '_', variant).strip('_')
        if v in thumbnails_on_disk:
            return f'{THUMB_DIR}/{v}.png'
    # Strategy 3: case-insensitive
    safe_lc = safe.lower()
    for t in thumbnails_on_disk:
        if t.lower() == safe_lc:
            return f'{THUMB_DIR}/{t}.png'
    return None

def make_thumbnail_image(thumb_path, px=THUMB_PX):
    """Resize the thumbnail in-memory and return an openpyxl Image."""
    src = PILImage.open(thumb_path).convert('RGB')
    src.thumbnail((px, px), PILImage.LANCZOS)
    canvas = PILImage.new('RGB', (px, px), (255, 255, 255))
    canvas.paste(src, ((px - src.width)//2, (px - src.height)//2))
    buf = io.BytesIO()
    canvas.save(buf, format='PNG')
    buf.seek(0)
    img = XLImage(buf)
    img.width = px
    img.height = px
    return img

# ============================================================
# Load workbook, insert column, embed thumbnails
# ============================================================
wb = load_workbook(XLSX)
ws = wb['Worksheet']

# Insert a new column at position 1 (pushes everything right)
ws.insert_cols(1)

# Set header for new column
hdr = ws.cell(row=1, column=1)
hdr.value = 'Creative'
hdr.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
hdr.fill = PatternFill('solid', fgColor='1F2A44')
hdr.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
hdr.border = BORDER

# Width of column A — wide enough for the thumbnail
# Excel column width unit is approximately 7 pixels per width unit
ws.column_dimensions['A'].width = 14   # ~98 px display

# Row heights to fit thumbnails (default ~15 = 20px; 65 ≈ 87px)
HEADER_HEIGHT = 28
ROW_HEIGHT = 65  # ≈ 87 pixels — fits 80px thumbnail with padding
ws.row_dimensions[1].height = HEADER_HEIGHT

embedded = 0
missing = []

for r in range(2, ws.max_row + 1):
    # Ad name is now in column 4 (was column 3, shifted right by insert_cols)
    ad_name = ws.cell(row=r, column=4).value
    ws.row_dimensions[r].height = ROW_HEIGHT

    # Border on the new cell
    cell = ws.cell(row=r, column=1)
    cell.border = BORDER
    cell.alignment = Alignment(horizontal='center', vertical='center')

    thumb_path = find_thumbnail(ad_name)
    if thumb_path:
        img = make_thumbnail_image(thumb_path, THUMB_PX)
        img.anchor = f'A{r}'
        ws.add_image(img)
        embedded += 1
    else:
        cell.value = '—'
        cell.font = Font(name='Calibri', size=9, italic=True, color='999999')
        missing.append(ad_name)

# Save
wb.save(XLSX)

print(f'✓ Saved {XLSX}')
print(f'  Embedded thumbnails: {embedded}')
print(f'  Missing thumbnails: {len(missing)}')
if missing:
    print(f'  Ads without thumbnails (new ads added after 26 May report):')
    for m in missing:
        print(f'    - {m}')
print(f'\n  File size: {os.path.getsize(XLSX)/1024:.0f} KB')
